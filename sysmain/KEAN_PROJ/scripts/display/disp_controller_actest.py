import openpyxl;
from openpyxl import *;
from openpyxl.utils import get_column_letter;
from openpyxl.styles import *;


import sys;
import os;
import ast;
import shutil;
import datetime;
from pathlib import Path;


sys.path.insert(0, str(Path(__file__).parents[1])+r'/utility');
import date_utils;

sys.path.insert(0, str(Path(__file__).parents[1])+r'/database');
import db_controller;




MONTH_DICT = {"June":"07","July":"08","August":"09","September":"10","October":"11","November":"12","December":"13","January":"02","Febuary":"03","March":"04","April":"05","May":"06"};

LOGO_PATH = '';



VALUE_SIGN_DICT = {'Energy Revenue':1,'Delivered Fuel Expense':-1,'Net Emissions Expense':-1,'Variable O&M Expense':-1,'Fixed Fuel':-1,'Hedge P&L':1,'Capacity Revenue':1,'Ancillary Services Revenue':1,'Misc Income':1,\
                    'Labor Expenses':-1,'Maintenance':-1,'Operations':-1,'Removal Costs':-1,'Fuel Handling':-1, 'Property Tax':-1, 'Insurance':-1, 'General & Administrative':-1,\
                    'Maintenance Capex':-1, 'Environmental Capex':-1, 'LTSA Capex':-1, 'Growth Capex':-1, 'Fixed Non-Labor Expense':-1, 'Total Fixed Costs':-1, 'Total Capex':-1}




def actest_report(conn_ins, company, amr_scenario, amr_version, budget_scenario, budget_version ,start_month=1, end_month=12):
    """ actest Report """
    ############################################################################################
    ############################################################################################
    print ("=================   generating amr report actuals + estimates...");



    start_month_str = str(start_month) if len(str(start_month)) > 1 else '0' + str(start_month);
    end_month_str = str(end_month) if len(str(end_month)) > 1 else '0' + str(start_month);

    report_date = str(datetime.datetime.now().date());

    year = int(amr_scenario.split(" ")[0]);

    report_start_date = str(year) + "-" + start_month_str + "-01";
    report_end_date = str(year) + "-" + end_month_str + "-31";
    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/bvr_report_template.xlsx';
    new_file_path = str(Path(__file__).parents[2]) + r'/reports/' + company + " " + amr_scenario + " " + amr_version + " " + report_date + ".xlsx";

    actest_result_df = db_controller.get_results_from_financials(conn_ins, company, amr_scenario, data_start_date=report_start_date, data_end_date = report_end_date, version = amr_version);
    budget_result_df = db_controller.get_results_from_financials(conn_ins, company, budget_scenario, data_start_date = report_start_date, data_end_date = report_end_date, version = budget_version);
    print (len(actest_result_df));
    print (len(budget_result_df));

    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';

    create_copy_of_template(ori_file_path, new_file_path);
    fill_in_cells_with_financials(new_file_path, actest_result_df, amr_scenario, budget_result_df);

    #############################################################################################
    #############################################################################################

    return new_file_path;












def create_copy_of_template(ori_file_path, new_file_path):
    shutil.copy(ori_file_path, new_file_path);



def fill_in_cells_with_financials(new_file_path, financials_result_df, scenario, budget_result_df):

    estimate_fill = PatternFill(fill_type='solid', start_color='FFE8E9EA', end_color='FFE8E9EA');

    grouped_by_entity_df = financials_result_df.groupby(['entity']);
    entity_name_list = list(grouped_by_entity_df.groups.keys());

    temp_entity_name_list = [];
    for item in entity_name_list:
        if item != 'HoldCo':
            temp_entity_name_list.append(item);

    temp_entity_name_list.append("HoldCo");


    entity_name_list = temp_entity_name_list; """ reorder the list, make sure holdco is the last tab """
    wb = load_workbook(filename = new_file_path);
    ws = wb.active;



    month_list = [];
    month_list_flag = 0;
    for selected_entity in entity_name_list:
        print (selected_entity);
        # selected_worksheet = wb.create_sheet(title=selected_entity);
        selected_worksheet = wb.copy_worksheet(ws);
        selected_worksheet.title = selected_entity;

        img = openpyxl.drawing.image.Image(LOGO_PATH)
        img.anchor(selected_worksheet['B1']);
        selected_worksheet.add_image(img);


        selected_worksheet.cell(row = 1, column = 13).value = selected_entity;
        selected_worksheet.cell(row = 1, column = 6).value = scenario;


        try:
            selected_financials_result_df = grouped_by_entity_df.get_group(selected_entity);
        except:
            for row_index in range(4,34):
                account = ws[start_col+str(row_index)].value;
                selected_worksheet.cell(row=row_index, column = 15).value = "=SUM(C" + str(row_index) + ":N" + str(row_index) + ")";
                temp_value = sum(list(budget_result_df.loc[(budget_result_df['account']==account) & (budget_result_df['entity'] == selected_entity)]['value']))/1000.0;
                if account in VALUE_SIGN_DICT:
                    temp_value = temp_value * VALUE_SIGN_DICT[account];
                selected_worksheet.cell(row=row_index, column = 16).value = temp_value;

            selected_worksheet.sheet_view.showGridLines = False;
            break;

        print (len(selected_financials_result_df));



        start_col = 'A';
        start_row = 4;
        end_row = 33;



        for row_index in range(start_row, end_row+1):
            account = ws[start_col+str(row_index)].value;
            account_value_df = selected_financials_result_df.loc[selected_financials_result_df['account'] == account];
            account_value_df = account_value_df.sort_values(['period']);
            row_value_list = [];
            # print (selected_worksheet.title, "    ",account, "    " , len(account_value_df));
            if month_list_flag == 0 and len(account_value_df) > 0:
                month_list = list(account_value_df['period']);
                # print (selected_entity);
                data_date_str = account_value_df.iloc[0]['scenario'];
                data_date = data_date_str.split(" ")[0] + "-" + MONTH_DICT[data_date_str.split(" ")[1]] + "-01"
                data_date = datetime.datetime.strptime(data_date, "%Y-%m-%d").date();
                month_list_flag = 1;

            for i in range(0, len(account_value_df)):
                # print (account_value_df.iloc[i]['period'], account_value_df.iloc[i]['value']);
                temp_value = account_value_df.iloc[i]['value']/1000.0;

                if account in VALUE_SIGN_DICT:
                    temp_value =temp_value * VALUE_SIGN_DICT[account];
                row_value_list.append(temp_value);


            for col_index in range(3,3+len(row_value_list)):
                list_index = col_index-3;
                selected_worksheet.cell(row=row_index, column = col_index).value = row_value_list[list_index];

        for col_index in range(3, 3+len(month_list)):
            list_index = col_index - 3;
            selected_worksheet.cell(row = 3, column=col_index).value = date_utils.calc_forecast_monthly_headers(month_list[list_index].year, month_list[list_index].month);
            if month_list[list_index] <= data_date:
                selected_worksheet.cell(row = 2, column=col_index).value = 'Act';
            else:
                selected_worksheet.cell(row = 2, column=col_index).value = 'Est';
                for fill_row in range(4,16):
                    selected_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;
                for fill_row in range(17,27):
                    selected_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;
                for fill_row in range(28,33):
                    selected_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;


        for row_index in range(4,34):
            account = ws[start_col+str(row_index)].value;
            selected_worksheet.cell(row=row_index, column = 15).value = "=SUM(C" + str(row_index) + ":N" + str(row_index) + ")";
            temp_value = sum(list(budget_result_df.loc[(budget_result_df['account']==account) & (budget_result_df['entity'] == selected_entity)]['value']))/1000.0;
            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];
            selected_worksheet.cell(row=row_index, column = 16).value = temp_value;




        selected_worksheet.sheet_view.showGridLines = False;


    first_worksheet = get_summary(ws, entity_name_list, scenario);



    # print (data_date);


    # print (month_list);

    for col_index in range(3, 3+len(month_list)):
        list_index = col_index - 3;
        if month_list[list_index] > data_date:
            for fill_row in range(4,16):
                first_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;
            for fill_row in range(17,27):
                first_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;
            for fill_row in range(28,33):
                first_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;

    if len(month_list) < 12:
        # print (range(3+len(month_list),3+12));
        for col_index in range(3+len(month_list)-1,3+12-1):
            for fill_row in range(4,16):
                first_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;
            for fill_row in range(17,27):
                first_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;
            for fill_row in range(28,33):
                first_worksheet.cell(row = fill_row, column=col_index).fill = estimate_fill;


    wb.save(new_file_path);




def get_summary(first_worksheet, entity_name_list, scenario):
    sum_string = "=SUM(";
    sum_string = sum_string + entity_name_list[0] + ":" + entity_name_list[-1] + "!"

    img = openpyxl.drawing.image.Image(LOGO_PATH)
    img.anchor(first_worksheet['B1']);
    first_worksheet.add_image(img);

    first_worksheet.title = 'Summary';

    first_worksheet.cell(row = 1, column = 6).value = scenario;
    first_worksheet.cell(row = 1, column = 13).value = 'Summary';

    for row_index in range(4,34):
        for col_index in range(3,18):
            first_worksheet.cell(row=row_index, column=col_index).value = sum_string + get_column_letter(col_index) + str(row_index) + ")";

    cell_string = "=" + entity_name_list[0] + "!";
    for col_index in range(3,15):
        first_worksheet.cell(row=2, column=col_index).value = cell_string + get_column_letter(col_index) + str(2) ;

    for col_index in range(3,18):
        first_worksheet.cell(row=3, column=col_index).value = cell_string + get_column_letter(col_index) + str(3) ;
    first_worksheet.sheet_view.showGridLines = False;

    return first_worksheet;
