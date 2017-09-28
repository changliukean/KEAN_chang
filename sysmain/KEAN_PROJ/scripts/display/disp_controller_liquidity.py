import openpyxl as opx
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

import openpyxl;
from openpyxl import *;
from openpyxl.utils import get_column_letter;
from openpyxl.styles import *;

import datetime;
import sys;
from pathlib import Path;
import shutil;


sys.path.insert(0, str(Path(__file__).parents[1])+r'\utility');
import date_utils;


LOGO_PATH = '';

def create_copy_of_template(ori_file_path, new_file_path):
    shutil.copy(ori_file_path, new_file_path);


def est_tax_dist_report(df_liquidity, tax_depreciation_result, scenario, company):
    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/est_tax_dist.xlsx';
    report_date = datetime.datetime.now().date();
    new_file_path = str(Path(__file__).parents[2]) + r'/reports/Est Tax Dist ' + company + ' ' + scenario + ' ' + str(report_date) + '.xlsx';

    create_copy_of_template(ori_file_path, new_file_path);

    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';



    fill_in_cells_tax_depreciation(new_file_path, tax_depreciation_result, scenario);
    fill_in_cells_est_tax_dist(new_file_path, df_liquidity, tax_depreciation_result, scenario);
    return new_file_path;


def fill_in_cells_est_tax_dist(new_file_path, df_liquidity, tax_depreciation_result, scenario):
    tax_wb = load_workbook(new_file_path);

    est_tax_dist = tax_wb['Estimated Tax Distribution'];
    current_year = int(scenario.split(" ")[0]);

    effective_date = 'Effective Date: ' + str(date_utils.get_month_dates_list(current_year, date_utils.get_month_number(scenario.split(" ")[1]), date_utils.get_month_number(scenario.split(" ")[1]))[0]);
    est_tax_dist['H1'] = effective_date;
    img = openpyxl.drawing.image.Image(LOGO_PATH)
    img.anchor(est_tax_dist['A1']);
    est_tax_dist.add_image(img);

    year_list = list(set(list(zip(*tax_depreciation_result))[0]));
    start_col = 2;
    # for item in tax_depreciation_result:
    #     print (item);
    for year in year_list:

        adj_ebitda = sum(list(df_liquidity.loc[df_liquidity.index.year == year]['EBITDA']));
        interest_expense = sum(list(df_liquidity.loc[df_liquidity.index.year == year]['Int Exp']));
        tax_depreciation = 0.0;
        if year == current_year:
            tax_depreciation = sum(list(zip(*[item for item in tax_depreciation_result if item[0]==year]))[6]);
        else:
            tax_depreciation = sum(list(zip(*[item for item in tax_depreciation_result if item[0]==year]))[-1]);

        est_tax_dist.cell(row = 5, column = start_col).value = adj_ebitda/1000;
        est_tax_dist.cell(row = 6, column = start_col).value = -interest_expense/1000;
        est_tax_dist.cell(row = 9, column = start_col).value = -tax_depreciation/1000;

        est_tax_dist.cell(row = 4, column = start_col).value = str(year);

        start_col += 2;



    tax_wb.save(new_file_path);


def fill_in_cells_tax_depreciation(new_file_path,tax_depreciation_result, scenario):

    # print (len(tax_depreciation_result));

    tax_wb = load_workbook(new_file_path);

    tax_depreciation_sheet = tax_wb['Tax Depreciation'];
    current_year = int(scenario.split(" ")[0]);

    effective_date = 'Effective Date: ' + str(date_utils.get_month_dates_list(current_year, date_utils.get_month_number(scenario.split(" ")[1]), date_utils.get_month_number(scenario.split(" ")[1]))[0]);
    tax_depreciation_sheet['L1'] = effective_date;
    img = openpyxl.drawing.image.Image(LOGO_PATH)
    img.anchor(tax_depreciation_sheet['A1']);
    tax_depreciation_sheet.add_image(img);


    for row_number in range(6,11):
        entity = tax_depreciation_sheet.cell(row=row_number, column=1).value;
        if entity == 'Goodwill':
            entity = 'HoldCo';
            year_list = list(set(sorted([item[0] for item in tax_depreciation_result])));
            tax_depreciation_sheet.cell(row = 4, column = 4).value = str(current_year);
            tax_depreciation_sheet.cell(row = 5, column = 4).value = 'Capex';
            tax_depreciation_sheet.cell(row = 5, column = 5).value = 'Total';
            tax_depreciation_sheet.cell(row = 4, column = 6).value = str(current_year)+' Tax';
            tax_depreciation_sheet.cell(row = 5, column = 6).value = 'Depreciation';
            year_column = 7;
            for year in year_list[1:]:
                tax_depreciation_sheet.cell(row = 4, column = year_column).value = str(year);
                tax_depreciation_sheet.cell(row = 5, column = year_column).value = 'Capex';
                year_column += 1;
                tax_depreciation_sheet.cell(row = 4, column = year_column).value = str(year) + ' Tax';
                tax_depreciation_sheet.cell(row = 5, column = year_column).value = 'Depreciation';
                year_column += 1;

        # print (entity);

        current_result_list = [item for item in tax_depreciation_result if item[1] == entity.strip()];

        # print (len(current_result_list));

        current_result_list = sorted(current_result_list, key=lambda x:x[0]);

        forecast_start_col = 7;
        for year_item in current_result_list:
            if year_item[0] == current_year:
                tax_depreciation_sheet.cell(row = row_number, column = 2).value = year_item[4]/1000;
                tax_depreciation_sheet.cell(row = row_number, column = 3).value = year_item[3];
                tax_depreciation_sheet.cell(row = row_number, column = 4).value = year_item[5]/1000;
                tax_depreciation_sheet.cell(row = row_number, column = 5).value = year_item[2]/1000;
                tax_depreciation_sheet.cell(row = row_number, column = 6).value = year_item[6]/1000;
            else:
                tax_depreciation_sheet.cell(row = row_number, column = forecast_start_col).value = year_item[2]/1000;
                forecast_start_col = forecast_start_col + 1;
                tax_depreciation_sheet.cell(row = row_number, column = forecast_start_col).value = year_item[-1]/1000;
                forecast_start_col = forecast_start_col + 1;



    tax_wb.save(new_file_path);







def liquidity_report(df_liquidity, scenario, company, maintenance_capex_dict, ltsa_capex_dict):
    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/Liquidity.xlsx';

    report_date = datetime.datetime.now().date();

    new_file_path = str(Path(__file__).parents[2]) + r'/reports/Liquidity ' + company + ' ' + scenario + ' ' + str(report_date) + '.xlsx';

    create_copy_of_template(ori_file_path, new_file_path);

    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';


    column_header_format = NamedStyle(name='datetime', number_format='mmm-yy')
    data_format = NamedStyle(name='data', number_format='#,##0.0_);(#,##0.0)')


    column_labels = df_liquidity[:datetime.date(2017,12,31)].index
    row_mapping_dict = {5:['Cash BOP', 1],6:['EBITDA', 1],7:['Capex', -1],8:['Change Work Cap', 1],
                9:['Other Cash Use', 1],10: ['Int Exp', -1],11:['Est Tax Dist',1],
                12: ['Revolver Change', 1],
                13: ['TLB Change', 1],14: ['DSRA Change', 1],15: ['Distributions', -1],
                16: ['Cash EOP', 1],19:['DSRA BOP', 1],20:['DSRA EOP', 1],21:['TLB BOP', 1],
                22:['TLB Addl Borrow',1], 23:['TLB Amort',-1], 24:['TLB Prepay',-1],
                25:['TLB EOP',1]};

    wb = load_workbook(new_file_path);
    ws_active = wb.active

    year_list = range(df_liquidity.index.min().date().year, df_liquidity.index.max().date().year+1);

    # print (year_list);

    # sys.exit();
    current_year = int(scenario.split(" ")[0]);

    effective_date = 'Effective Date: ' + str(date_utils.get_month_dates_list(current_year, date_utils.get_month_number(scenario.split(" ")[1]), date_utils.get_month_number(scenario.split(" ")[1]))[0]);

    for year in year_list:

        selected_worksheet = wb.copy_worksheet(wb['year']);
        selected_worksheet.title = str(year);
        dft = df_liquidity[datetime.date(year, 1,1):datetime.date(year, 12,31)];

        # ws.column_dimensions['A'].width=22.0

        img = openpyxl.drawing.image.Image(LOGO_PATH)
        img.anchor(selected_worksheet['A1']);
        selected_worksheet.add_image(img);

        selected_worksheet['K1'] = effective_date;

        if year == current_year:
            for col_number in range(2,14):
                if col_number-1 < date_utils.get_month_number(scenario.split(" ")[1]):
                    selected_worksheet.cell(row = 3, column = col_number).value = 'Act';
                else:
                    selected_worksheet.cell(row = 3, column = col_number).value = 'Est';
        else:
            for col_number in range(2,14):
                selected_worksheet.cell(row = 3, column = col_number).value = 'Forecast';

        for col_number in range(2,14):
            selected_worksheet.cell(row = 4, column = col_number).value = date_utils.calc_forecast_monthly_headers(list(dft.index)[col_number-2].date().year, list(dft.index)[col_number-2].date().month);

        for row_number in range(5,17):
            for col_number in range(2,14):
                selected_worksheet.cell(row = row_number, column = col_number).value = dft.iloc[col_number-2][row_mapping_dict[row_number][0]]*row_mapping_dict[row_number][1]/1000000.0;

        for row_number in range(19,26):
            for col_number in range(2,14):
                selected_worksheet.cell(row = row_number, column = col_number).value = dft.iloc[col_number-2][row_mapping_dict[row_number][0]]*row_mapping_dict[row_number][1]/1000000.0;

        for row_number in range(26,28):
            for col_number in range(2,14):
                selected_worksheet.cell(row = row_number, column = col_number).value = 0;

        maintenance_capex_list = maintenance_capex_dict[year];
        ltsa_capex_list = ltsa_capex_dict[year];
        maintenance_capex_row_number = 34;
        ltsa_capex_row_number = 35;
        for col_number in range(2,14):
            selected_worksheet.cell(row = maintenance_capex_row_number, column = col_number).value = maintenance_capex_list[col_number-2]/1000000.0;
            selected_worksheet.cell(row = ltsa_capex_row_number, column = col_number).value = ltsa_capex_list[col_number-2]/1000000.0;


    wb = create_summary_liquidity(wb, scenario, company);


    wb.remove_sheet(wb['year']);




    wb.save(new_file_path);
    return new_file_path;


def create_summary_liquidity(workbook, scenario, company):
    current_year = int(scenario.split(" ")[0]);



    summary_sheet = workbook['Summary'];
    summary_sheet['B1'] = 'Debt Compliance and Liquidity';
    summary_sheet['J1'] = scenario;
    summary_sheet['A5'] = company + " liquidity forecast";
    summary_sheet['A22'] = company + " debt covenant forecast";


    for col_number in range(3,13,3):
        month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
        summary_sheet.cell(row = 5, column = col_number).value = month_str;

    for row_number in range(6,7):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=\'" + str(current_year) + "\'!" + get_column_letter(col_number-1)+"5";

    for row_number in range(7,12):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=SUM(\'" + str(current_year) + "\'!" + get_column_letter(col_number-1) + str(row_number-1) + ":" + get_column_letter(col_number+1) + str(row_number-1) + ")";

    for row_number in range(12,16):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=SUM(\'" + str(current_year) + "\'!" + get_column_letter(col_number-1) + str(row_number) + ":" + get_column_letter(col_number+1) + str(row_number) + ")";



    for row_number in range(16,17):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=\'" + str(current_year) + "\'!" + get_column_letter(col_number+1)+"16";


    for row_number in range(18,19):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=" + get_column_letter(col_number)+"16";

    for row_number in range(19,20):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = 99950000/1000000;

    for row_number in range(21,22):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=\'" + str(current_year) + "\'!" + get_column_letter(col_number+1)+"20";



    for row_number in range(23,24):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=\'" + str(current_year) + "\'!" + get_column_letter(col_number-1)+"21";


    for row_number in range(24,27):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value =  "=SUM(\'" + str(current_year) + "\'!" + get_column_letter(col_number-1) + str(row_number-2) + ":" + get_column_letter(col_number+1) + str(row_number-2) + ")";


    for row_number in range(34,35):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = str(int(current_year + col_number/3));

    for row_number in range(35,36):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=-SUM(\'" + str(int(current_year + col_number/3)) + "\'!B10:M10)-SUM(\'" + str(int(current_year + col_number/3)) + "\'!B23:M23)";

    for row_number in range(36,37):
        for col_number in range(3,13,3):
            month_str = date_utils.calc_forecast_monthly_headers(current_year, col_number);
            summary_sheet.cell(row = row_number, column = col_number).value = "=(SUM(\'" + str(int(current_year + col_number/3)) + "\'!B6:M6)-SUM(\'" + str(int(current_year + col_number/3)) + "\'!B36:M36))/Summary!" + get_column_letter(col_number) + "35";



    return workbook;





def debt_balance_report(df_liquidity, scenario, company):
    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/debt_balances.xlsx';

    report_date = datetime.datetime.now().date();

    new_file_path = str(Path(__file__).parents[2]) + r'/reports/Debt Balance ' + company + " " + scenario + " " + str(report_date) + ".xlsx";


    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';

    create_copy_of_template(ori_file_path, new_file_path);
    fill_in_cells_debt_balances(new_file_path, df_liquidity, scenario);
    return new_file_path;



def interest_expense_report(df_liquidity, scenario, company):
    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/interest_expense.xlsx';
    report_date = datetime.datetime.now().date();
    new_file_path = str(Path(__file__).parents[2]) + r'/reports/Interest Expense ' + company + " " + scenario + " " + str(report_date) + ".xlsx";
    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';

    create_copy_of_template(ori_file_path, new_file_path);
    fill_in_cells_interest_expense(new_file_path, df_liquidity, scenario);
    return new_file_path;



def fill_in_cells_interest_expense(new_file_path, df_liquidity, scenario):
    current_year = int(scenario.split(" ")[0]);
    year_list = range(df_liquidity.index.min().date().year, df_liquidity.index.max().date().year+1);

    wb = load_workbook(new_file_path);
    ws_active = wb.active;
    effective_date = 'Effective Date: ' + str(date_utils.get_month_dates_list(current_year, date_utils.get_month_number(scenario.split(" ")[1]), date_utils.get_month_number(scenario.split(" ")[1]))[0]);

    for year in year_list:
        selected_worksheet = wb.copy_worksheet(ws_active);
        selected_worksheet.title = str(year);
        img = openpyxl.drawing.image.Image(LOGO_PATH)
        img.anchor(selected_worksheet['A1']);
        selected_worksheet.add_image(img);
        selected_worksheet['N1'] = effective_date;
        dft = df_liquidity[datetime.date(year, 1,1):datetime.date(year, 12,31)];

        row_mapping_dict = {6:['TLB BOP',1], 7:['TLB Float Rate',1000000], 8:['Swap Avg Daily Balance',1],
                            9:['Swap Fix Rate',1000000], 10:['TLB Int Exp',1],
                            15:['TLC BOP',1],16:['TLC Float Rate',1000000],17:['TLC Int Exp',1],
                            22:['TLB Float Rate',1000000]}

        if year == current_year:
            for col_number in range(2,14):
                if col_number-1 <= date_utils.get_month_number(scenario.split(" ")[1]):
                    selected_worksheet.cell(row = 3, column = col_number).value = 'Act';
                else:
                    selected_worksheet.cell(row = 3, column = col_number).value = 'Est';
        else:
            for col_number in range(2,14):
                selected_worksheet.cell(row = 3, column = col_number).value = 'Forecast';

        # column_header_format = NamedStyle(name='datetime', number_format='mmm-yy')

        for col_number in range(2,14):
            temp_date_str = str(list(dft.index)[col_number-2].date());

            # day_str = temp_date_str.split("-")[1] if '0' != temp_date_str.split("-")[1][0] else temp_date_str.split("-")[1][1:];
            """
                temp_date_str.split("-")[1] + "/" + temp_date_str.split("-")[2] + "/"+ temp_date_str.split("-")[0]
            """

            selected_worksheet.cell(row = 4, column = col_number).value = list(dft.index)[col_number-2].date();
            selected_worksheet.cell(row = 4, column = col_number).number_format = 'mmm-yy';
            # selected_worksheet.cell(row = 4, column = col_number).style = column_header_format;

        for row_number in range(6,11):
            for col_number in range(2,14):
                selected_worksheet.cell(row = row_number, column = col_number).value = dft.iloc[col_number-2][row_mapping_dict[row_number][0]]*row_mapping_dict[row_number][1]/1000000.0;

        for col_number in range(2,14):
            effective_interest_rate = 0.0;
            days = 2 if list(dft.index)[col_number-2].date() == datetime.datetime(2017,1,31).date() else list(dft.index)[col_number-2].date().day;
            effective_interest_rate = dft.iloc[col_number-2]['TLB Int Exp'] / dft.iloc[col_number-2]['TLB BOP'] * 360 / days;
            selected_worksheet.cell(row = 11, column = col_number).value = effective_interest_rate;


        for row_number in range(15,18):
            for col_number in range(2,14):
                selected_worksheet.cell(row = row_number, column = col_number).value = dft.iloc[col_number-2][row_mapping_dict[row_number][0]]*row_mapping_dict[row_number][1]/1000000.0;

        for row_number in range(21,24):
            for col_number in range(2,14):
                if row_number in [21,23]:
                    selected_worksheet.cell(row = row_number, column = col_number).value = 0.0;
                else:
                    selected_worksheet.cell(row = row_number, column = col_number).value = dft.iloc[col_number-2][row_mapping_dict[row_number][0]]*row_mapping_dict[row_number][1]/1000000.0 - 0.01;


        selected_worksheet['N4'] = 'FY ' + str(year);

        day_list = [day for day in [int(day_item.date().day) for day_item in list(dft.index)]];


        if year == 2017:
            day_list[0] = 2;
            # total_tlb_bop = sum([tlb_bop * day for tlb_bop in list(dft['TLB BOP']) for day in [int(day_item.date().day) for day_item in list(dft.index)]]);


        tlb_bop_list = [tlb_bop for tlb_bop in list(dft['TLB BOP'])];
        # print (len(tlb_bop_list), len(day_list));
        total_tlb_bop = sum([tlb_bop_list[i] * day_list[i] for i in range(0,12)]);



        # print (year, (total_tlb_bop)/1000000);
        # print (sum(day_list));
        selected_worksheet['N6'] = total_tlb_bop/sum(day_list)/1000000;

        float_rate_list = list(dft['TLB Float Rate']);

        total_tlb_float_rate = sum([tlb_bop_list[i] * float_rate_list[i] * day_list[i] for i in range(0,12)])

        selected_worksheet['N7'] = total_tlb_float_rate/sum(day_list)/1000000/(total_tlb_bop/sum(day_list)/1000000);

        swap_avg_list = list(dft['Swap Avg Daily Balance']);

        total_swap_avg = sum([swap_avg_list[i] * day_list[i] for i in range(0, 12) if swap_avg_list[i] > 0]);

        try:
            selected_worksheet['N8'] = total_swap_avg/sum([day_list[i] for i in range(0,12) if swap_avg_list[i] > 0])/1000000;
        except:
            selected_worksheet['N8'] = 0;

        swap_fix_rate_list = list(dft['Swap Fix Rate']);

        total_fix_interest = sum([swap_fix_rate_list[i] * swap_avg_list[i]/1000000 * day_list[i] for i in range(0, 12) if swap_avg_list[i] > 0]);

        try:
            selected_worksheet['N9'] = total_fix_interest/sum([day_list[i] for i in range(0,12) if swap_avg_list[i] > 0])/(total_swap_avg/sum([day_list[i] for i in range(0,12) if swap_avg_list[i] > 0])/1000000);
        except:
            selected_worksheet['N9'] = 0;

        selected_worksheet['N10'] = sum(list(dft['TLB Int Exp']))/1000000;

        selected_worksheet['N11'] = sum(list(dft['TLB Int Exp']))/(total_tlb_bop/sum(day_list)) * 360 / sum(day_list) ;


        tlc_bop_list = [tlc_bop for tlc_bop in list(dft['TLC BOP'])];
        total_tlc_bop = sum([tlc_bop_list[i] * day_list[i] for i in range(0, 12)]);

        selected_worksheet['N15'] = total_tlc_bop/sum(day_list)/1000000;

        float_rate_list = list(dft['TLC Float Rate']);

        total_tlc_float_rate = sum([tlc_bop_list[i] * float_rate_list[i] * day_list[i] for i in range(0,12)])


        selected_worksheet['N16'] = total_tlc_float_rate/sum(day_list)/1000000/(total_tlc_bop/sum(day_list)/1000000);
        selected_worksheet['N17'] = sum(list(dft['TLC Int Exp']))/1000000;



        selected_worksheet['N21'] = 0;
        selected_worksheet['N22'] = 0;
        selected_worksheet['N23'] = 0;



    wb.remove_sheet(ws_active);
    wb.save(new_file_path);




def fill_in_cells_debt_balances(new_file_path, df_liquidity, scenario):
    current_year = int(scenario.split(" ")[0]);
    year_list = range(df_liquidity.index.min().date().year, df_liquidity.index.max().date().year+1);

    effective_date = 'Effective Date: ' + str(date_utils.get_month_dates_list(current_year, date_utils.get_month_number(scenario.split(" ")[1]), date_utils.get_month_number(scenario.split(" ")[1]))[0]);


    wb = load_workbook(filename = new_file_path);
    ws_active = wb.active;

    for year in year_list:
        selected_worksheet = wb.copy_worksheet(ws_active);
        selected_worksheet.title = str(year);
        img = openpyxl.drawing.image.Image(LOGO_PATH)
        img.anchor(selected_worksheet['A1']);
        selected_worksheet.add_image(img);

        selected_worksheet['M1'] = effective_date;

        dft = df_liquidity[datetime.date(year, 1,1):datetime.date(year, 12,31)];


        if year == current_year:
            for col_number in range(2,14):
                if col_number-1 <= date_utils.get_month_number(scenario.split(" ")[1]):
                    selected_worksheet.cell(row = 3, column = col_number).value = 'Act';
                else:
                    selected_worksheet.cell(row = 3, column = col_number).value = 'Est';
        else:
            for col_number in range(2,14):
                selected_worksheet.cell(row = 3, column = col_number).value = 'Forecast';

        for col_number in range(2,14):
            selected_worksheet.cell(row = 4, column = col_number).value = date_utils.calc_forecast_monthly_headers(list(dft.index)[col_number-2].date().year, list(dft.index)[col_number-2].date().month);

        """
            TLB Addl Borrow	TLB Amort	TLB BOP	TLB Change	TLB EOP	TLB Float Rate	TLB Int Exp	TLB Prepay
            TLC Addl Borrow	TLC Amort	TLC BOP	TLC EOP	TLC Float Rate	TLC Int Exp	TLC Prepay
        """

        row_mapping_dict = {6:['TLB BOP',1], 7:['TLB Addl Borrow',1], 8:['TLB Amort',-1], 9:['TLB Prepay',-1], 10:['TLB EOP',1],\
                            14:['TLC BOP',1],15:['TLC Prepay',-1],16:['TLC EOP',1],\
                            20:[],21:[],22:[],23:[]}
        for row_number in range(6,11):
            for col_number in range(2,14):
                selected_worksheet.cell(row = row_number, column = col_number).value = dft.iloc[col_number-2][row_mapping_dict[row_number][0]]*row_mapping_dict[row_number][1]/1000000.0;

        for row_number in range(14,17):
            for col_number in range(2,14):
                selected_worksheet.cell(row = row_number, column = col_number).value = dft.iloc[col_number-2][row_mapping_dict[row_number][0]]*row_mapping_dict[row_number][1]/1000000.0;

        for row_number in range(20,24):
            for col_number in range(2,14):
                selected_worksheet.cell(row = row_number, column = col_number).value = 0;


    wb.remove_sheet(wb.active);
    wb.save(new_file_path);
