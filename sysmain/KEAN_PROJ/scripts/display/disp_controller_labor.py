import xlsxwriter;
from openpyxl import *;
from openpyxl.utils import get_column_letter;
from openpyxl.writer.write_only import WriteOnlyCell;
from openpyxl.comments import Comment;
from openpyxl.styles import *;
import openpyxl;

import sys;
import os;
import ast;
import shutil;
import datetime;
from pathlib import Path;


sys.path.insert(0, str(Path(__file__).parents[1])+r'/utility');
import date_utils;




COMPANY_ENTITY_MAP = {'Lightstone':['Gavin','Lawrenceburg','Waterford','Darby']}




def create_copy_of_template(ori_file_path, new_file_path):
    shutil.copy(ori_file_path, new_file_path);



def labor_report(support_document_labor_expense_list, scenario, company):
    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/labor_expense.xlsx';
    report_date = datetime.datetime.now().date();
    new_file_path = str(Path(__file__).parents[2]) + r'/reports/Labor Expense ' + company + ' ' + scenario + ' ' + str(report_date) + '.xlsx';
    create_copy_of_template(ori_file_path, new_file_path);



    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';
    fill_in_cells_labor(new_file_path, support_document_labor_expense_list, scenario, company);



def fill_in_cells_labor(new_file_path, support_document_labor_expense_list, scenario, company):

    actuals_month_dates_list, allyear_month_dates_list, estimate_month_dates_list = date_utils.get_dates_info_from_amr_scenario(scenario);

    entity_name_list = COMPANY_ENTITY_MAP[company];

    wb = load_workbook(new_file_path);

    for entity in entity_name_list:
        selected_worksheet = wb.copy_worksheet(wb.active);
        selected_worksheet.title = entity;
        selected_worksheet['K1'] = scenario;


        entity_support_list = [item for item in support_document_labor_expense_list if item[0] == entity];
        print (len(entity_support_list));

        for i in range(0, len(allyear_month_dates_list)):
            period = allyear_month_dates_list[i];
            selected_worksheet.cell(row = 3, column = i+2).value = date_utils.calc_forecast_monthly_headers(int(period.split("-")[0]), int(period.split("-")[1]));

            img = openpyxl.drawing.image.Image(LOGO_PATH)
            img.anchor(selected_worksheet['A1']);
            selected_worksheet.add_image(img);

            # print (period);
            if period in actuals_month_dates_list:
                continue;
            else:
                period_entity_support_list = [item for item in entity_support_list if item[1] == period][0];
                """
                    [entity, period, basesalary, incentive, fringe, overtime, ldw_credit, severence, retention, retiree_medical, headcount, payroll_cycles]
                """
                selected_worksheet.cell(row = 5, column = i+2).value = period_entity_support_list[-2];
                selected_worksheet.cell(row = 6, column = i+2).value = period_entity_support_list[-1];
                selected_worksheet.cell(row = 9, column = i+2).value = period_entity_support_list[2];
                selected_worksheet.cell(row = 10, column = i+2).value = period_entity_support_list[3];
                selected_worksheet.cell(row = 11, column = i+2).value = period_entity_support_list[4];
                selected_worksheet.cell(row = 12, column = i+2).value = period_entity_support_list[5];
                selected_worksheet.cell(row = 15, column = i+2).value = period_entity_support_list[6];
                selected_worksheet.cell(row = 16, column = i+2).value = period_entity_support_list[7];
                selected_worksheet.cell(row = 17, column = i+2).value = period_entity_support_list[8];
                selected_worksheet.cell(row = 18, column = i+2).value = period_entity_support_list[9];


    wb.remove_sheet(wb.active);

    wb.save(new_file_path);
