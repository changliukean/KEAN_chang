import openpyxl as opx
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

import openpyxl;
from openpyxl import *;
from openpyxl.utils import get_column_letter;
from openpyxl.styles import *;

import datetime;
import sys;
from pathlib import Path;
import shutil;


sys.path.insert(0, str(Path(__file__).parents[1])+r'/utility');
import date_utils;


sys.path.insert(0, str(Path(__file__).parents[1])+r'/logic');
from hedge import get_specific_price, calc_monthly_cashflow;




from datetime import date;

LOGO_PATH = '';



def create_copy_of_template(ori_file_path, new_file_path):
    shutil.copy(ori_file_path, new_file_path);


def hedge_report(df_hedges, df_all_prices, df_hours_pjm ,scenario, company):
    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/hedges.xlsx';
    report_date = datetime.datetime.now().date();
    new_file_path = str(Path(__file__).parents[2]) + r'/reports/Hedges ' + company + ' ' + scenario + ' ' + str(report_date) + '.xlsx';

    create_copy_of_template(ori_file_path, new_file_path);

    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';
    fill_in_cells_hedges(new_file_path, df_hedges, df_all_prices, df_hours_pjm ,scenario);



def fill_in_cells_hedges(new_file_path,df_hedges, df_all_prices , df_hours_pjm ,scenario):
    wb = load_workbook(new_file_path);
    ws = wb['hedges'];


    img = openpyxl.drawing.image.Image(LOGO_PATH)
    img.anchor(ws['A1']);
    ws.add_image(img);

    forecast_start = date(int(scenario.split(" ")[0]), date_utils.get_month_number(scenario.split(" ")[1])+1,1);
    # print (forecast_start);

    # month_header_list = date_utils.get_month_dates_list(forecast_start.year, forecast_start.month, 12);
    # print (month_header_list);

    start_col = 9;
    for item in range(forecast_start.month,13):
        ws.cell(row = 4, column = start_col).value = date_utils.calc_forecast_monthly_headers(forecast_start.year,item);
        start_col += 1;

    ws.cell(row=4, column=start_col).value = '2018+';

    forecast_end = date(int(scenario.split(" ")[0])+4, 12,31);
    # print (forecast_end);
    valuation_date = date(int(scenario.split(" ")[0]), date_utils.get_month_number(scenario.split(" ")[1]),28)

    loop_rows = 1
    remaining_column = 8 + (13 - forecast_start.month) + 1
    data_format = NamedStyle(name='data', number_format='#,##0_);(#,##0)')
    df_hedges_relevant = df_hedges[df_hedges['trade_date']<forecast_start]
    df_hedges_relevant = df_hedges_relevant.reset_index(drop=True)


    for i, hedge in df_hedges_relevant.iterrows():
        ws.cell(row=i+5, column=1, value=hedge.loc['tid_kindle'])
        ws.cell(row=i+5, column=2, value=hedge.loc['counterparty'])
        instrument = hedge.loc['receive_index'] + "/" + hedge.loc['pay_index']
        ws.cell(row=i+5, column=3, value=instrument)
        ws.cell(row=i+5, column=4, value=hedge.loc['notional'])
        fixed_price = hedge.loc['receive_value'] if hedge.loc['receive_index'] == 'fixed' else hedge.loc['pay_value']
        ws.cell(row=i+5, column=5, value=fixed_price)
        ws.cell(row=i+5, column=6, value=hedge.loc['frequency'])
        ws.cell(row=i+5, column=7, value=hedge.loc['date_start'])
        ws.cell(row=i+5, column=8, value=hedge.loc['date_end'])


        """show monthly results for remainder of forecast year, then sum remaining into one column"""
        loop_date = date_utils.calc_month_end(forecast_start, 'date')
        month_offset = 0
        remaining_cashflow = 0

        while loop_date <= forecast_end:
            if hedge.loc['date_start'] <= loop_date and hedge.loc['date_end'] >= loop_date:
                price_receive = hedge.loc['receive_value'] if hedge.loc['receive_index'] == 'fixed' else get_specific_price(df_all_prices, hedge.loc['receive_index'], loop_date)
                price_pay = hedge.loc['pay_value'] if hedge.loc['pay_index'] == 'fixed' else get_specific_price(df_all_prices, hedge.loc['pay_index'], loop_date)
                monthly_cashflow = 0.0;
                if hedge.loc['pay_index'] == 'AD Hub DA Peak':
                    monthly_cashflow = calc_monthly_cashflow(price_receive, price_pay, hedge.loc['notional'], hedge.loc['frequency'], loop_date, df_hours_pjm, option = True)
                else:
                    monthly_cashflow = calc_monthly_cashflow(price_receive, price_pay, hedge.loc['notional'], hedge.loc['frequency'], loop_date, df_hours_pjm)

                if loop_date.year == forecast_start.year:
                    ws.cell(row=i+5, column=9+month_offset, value = monthly_cashflow)
                    ws.cell(row=i+5, column=9+month_offset).style = data_format
                else:
                    remaining_cashflow += monthly_cashflow
            if loop_date > hedge.loc['date_end']:
                ws.cell(row=i+5, column=remaining_column, value = remaining_cashflow)
                ws.cell(row=i+5, column=remaining_column).style = data_format
                break

            month_offset += 1
            loop_date = date_utils.calc_next_month_end(loop_date, 'date')

        loop_rows += 1


    ''' add totals row '''
    ws.cell(row=loop_rows+4, column=1, value='Total')
    for i in range(9, remaining_column+1):
        formula = '=SUM(' + chr(96+i) + str(1+4) + ':' + chr(96+i) + str(loop_rows+4-1) + ')'
        ws.cell(row=loop_rows+4, column=i, value=formula)
        ws.cell(row=loop_rows+4, column=i).style = data_format

    ''' Add bold font and grey fill to total row '''
    for i in range(1, remaining_column+1):
        current_cell = chr(96+i) + str(loop_rows+4)
        ws[current_cell].fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
        ws[current_cell].font = Font(bold=True)

    ''' Add underline to report title '''
    for i in range(1, remaining_column+1):
        current_cell = chr(96+i) + str(1)
        ws[current_cell].border = Border(bottom=Side(style='thick'))

    label = 'Pricing Date: ' +  valuation_date.strftime('%m/%d/%Y')
    ws.cell(row=1, column=remaining_column-1, value=scenario)
    ws.cell(row=loop_rows+7,column=1,value=label);

    wb.save(new_file_path);
