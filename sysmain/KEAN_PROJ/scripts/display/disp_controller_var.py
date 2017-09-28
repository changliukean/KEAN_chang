import openpyxl;
from openpyxl import *;
from openpyxl.utils import get_column_letter;
from openpyxl.styles import *;


import sys;
import os;
import ast;
import shutil;
import datetime;
from pathlib import Path;


sys.path.insert(0, str(Path(__file__).parents[1])+r'/utility');
import date_utils;

sys.path.insert(0, str(Path(__file__).parents[1])+r'/database');
import db_controller;


YTD_OPTION  = 'YTD';
GAS_PLANTS_LIST = ['Waterford', 'Lawrenceburg', 'Darby'];



VALUE_SIGN_DICT = {'Energy Revenue':1,'Delivered Fuel Expense':-1,'Net Emissions Expense':-1,'Variable O&M Expense':-1,'Fixed Fuel':-1,'Hedge P&L':1,'Capacity Revenue':1,'Ancillary Services Revenue':1,'Misc Income':1,\
                    'Labor Expenses':-1,'Maintenance':-1,'Operations':-1,'Removal Costs':-1,'Fuel Handling':-1, 'Property Tax':-1, 'Insurance':-1, 'General & Administrative':-1,\
                    'Maintenance Capex':-1, 'Environmental Capex':-1, 'LTSA Capex':-1, 'Growth Capex':-1, 'Fixed Non-Labor Expense':-1, 'Total Fixed Costs':-1, 'Total Capex':-1}




def variance_report(conn_ins, company_name, amr_scenario, amr_version, budget_scenario, budget_version, actuals_month_dates_list):
    """ year_to_date and current month  BVReports """
    ###########################################################################################
    ###########################################################################################
    print ("=================   generating amr variance report ytd & mtd...");
    report_date = str(datetime.datetime.now().date());


    report_start_date = actuals_month_dates_list[0];
    report_end_date = actuals_month_dates_list[-1];


    # print (report_date);
    # print (report_start_date);
    # print (report_end_date);

    ytd_actest_result_df = db_controller.get_results_from_financials(conn_ins, company_name, amr_scenario, data_start_date=report_start_date, data_end_date = report_end_date, version = amr_version);
    ytd_budget_result_df = db_controller.get_results_from_financials(conn_ins, company_name, budget_scenario, data_start_date = report_start_date, data_end_date = report_end_date, version = budget_version);


    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/bvr_report_template_variance.xlsx';
    new_file_path_ytd = str(Path(__file__).parents[2]) + r'/reports/' + company_name + " " + amr_scenario + " variance ytd " + report_date + ".xlsx";
    new_file_path_month = str(Path(__file__).parents[2]) + r'/reports/' + company_name + " " + amr_scenario + " variance mtd " + report_date + ".xlsx";
    create_copy_of_template(ori_file_path, new_file_path_ytd);
    create_copy_of_template(ori_file_path, new_file_path_month);
    ytd_date = report_end_date;
    fill_in_cells_with_financials_ytd(new_file_path_ytd, ytd_actest_result_df, ytd_budget_result_df, ytd_date);

    mtd_day = sorted(set(list(ytd_actest_result_df['period'])))[-1];


    # print (mtd_day);
    month_actest_result_df = ytd_actest_result_df.loc[ytd_actest_result_df['period'] == mtd_day];
    month_budget_result_df = ytd_budget_result_df.loc[ytd_budget_result_df['period'] == mtd_day];


    fill_in_cells_with_financials_month(new_file_path_month, month_actest_result_df, month_budget_result_df, ytd_date);

    ############################################################################################
    ###########################################################################################
    return new_file_path_ytd, new_file_path_month;



def create_copy_of_template(ori_file_path, new_file_path):
    shutil.copy(ori_file_path, new_file_path);


def fill_in_cells_with_financials_ytd(new_file_path, actest_result_df, budget_result_df, ytd_date, holdco_page = True):

    month = int(ytd_date.split("-")[1]);
    month_header = date_utils.calc_forecast_monthly_headers_month(month);
    # print (month_header+" YTD");

    workbook = load_workbook(new_file_path);

    summary_worksheet = workbook.get_sheet_by_name('summary');

    per_entity_worksheet = workbook.get_sheet_by_name('per_entity');

    holdco_worksheet = workbook.get_sheet_by_name('holdco');



    # ytd_gas_plants_worksheet = workbook.copy_worksheet(ytd_worksheet);
    # ytd_gas_plants_worksheet.title = "Gas Plants " + month_header + " YTD";

    ytd_summary_worksheet = workbook.copy_worksheet(summary_worksheet);
    ytd_summary_worksheet.title = "Summary " + month_header + " " + YTD_OPTION;

    draw_var_ytd(month_header, actest_result_df, budget_result_df, ytd_summary_worksheet, option='summary');

    ytd_gas_plants_worksheet = workbook.copy_worksheet(summary_worksheet);
    ytd_gas_plants_worksheet.title = "Gas Plants " + month_header + " " + YTD_OPTION;

    gas_plants_actest_result_df = actest_result_df.loc[actest_result_df['entity'].isin(GAS_PLANTS_LIST)];
    gas_plants_budget_result_df = budget_result_df.loc[budget_result_df['entity'].isin(GAS_PLANTS_LIST)];

    draw_var_ytd(month_header, gas_plants_actest_result_df, gas_plants_budget_result_df, ytd_gas_plants_worksheet, option='gas_plants');


    grouped_actest_df = actest_result_df.groupby(['entity']);

    for entity in grouped_actest_df.groups:
        entity_worksheet = workbook.copy_worksheet(per_entity_worksheet);
        entity_worksheet.title = entity + " " + month_header + " " + YTD_OPTION;
        entity_actest_df = grouped_actest_df.get_group(entity);
        entity_budget_df = budget_result_df.loc[(budget_result_df['entity'] == entity)];
        draw_var_per_entity(month_header, entity_actest_df, entity_budget_df, entity_worksheet, entity);


    if holdco_page:
        selected_holdco_worksheet = workbook.copy_worksheet(holdco_worksheet);
        selected_holdco_worksheet.title = 'HoldCo YTD Detail';

        draw_var_holdco(month_header, actest_result_df, budget_result_df, selected_holdco_worksheet);


    workbook.remove_sheet(workbook.get_sheet_by_name('summary'));
    workbook.remove_sheet(workbook.get_sheet_by_name('per_entity'));
    workbook.remove_sheet(workbook.get_sheet_by_name('holdco'));

    workbook.save(new_file_path);



def draw_var_holdco(month_header, actest_result_df, budget_result_df, selected_worksheet):
    selected_worksheet.sheet_view.showGridLines = False;


    holdco_fsli_item_dict = {'Lightstone': ['Hedge P&L','Arctos','Kindle Energy (AMA fees)',
                                'Nextera (EMA fees)','CAMS (OMA fees)','Transition Costs','General & Administrative']}

    current_company = actest_result_df.iloc[0]['company'];
    current_scenario = actest_result_df.iloc[0]['scenario'];


    current_month_date = date_utils.get_dates_info_from_amr_scenario(current_scenario)[0][-1];

    print (current_month_date);

    current_month_date_obj = datetime.date(int(current_month_date.split("-")[0]), int(current_month_date.split("-")[1]), int(current_month_date.split("-")[2]));

    fsli_item_list = holdco_fsli_item_dict[current_company];

    holdco_actual_result_df = actest_result_df.loc[(actest_result_df['account'].isin(fsli_item_list)) & (actest_result_df['entity'] == 'HoldCo')];

    holdco_budget_result_df = budget_result_df.loc[(budget_result_df['account'].isin(fsli_item_list)) & (budget_result_df['entity'] == 'HoldCo')];



    # for i in range(0, len(holdco_actual_result_df)):
    #     print (holdco_actual_result_df.iloc[i]);
    #
    # for i in range(0, len(holdco_budget_result_df)):
    #     print (holdco_budget_result_df.iloc[i]);

    # print (month_header);




    for row_number in range(2, 10):
        if row_number == 2:
            selected_worksheet.cell(row = row_number, column = 3).value = "HoldCo " + month_header + " MTD" ;
            selected_worksheet.cell(row = row_number, column = 6).value = "HoldCo " + month_header + " YTD" ;
            continue;
        if row_number == 3:
            selected_worksheet.cell(row = row_number, column = 3).value = "Actual";
            selected_worksheet.cell(row = row_number, column = 4).value = 'Budget';
            selected_worksheet.cell(row = row_number, column = 5).value = 'Variance';
            continue;
        account_title = selected_worksheet.cell(row = row_number, column = 1).value;
        # print ("-----------------------", account_title);
        if account_title != 'Hedge P&L':
            selected_worksheet.cell(row = row_number, column = 3).value = holdco_actual_result_df.loc[(holdco_actual_result_df['entity'] == 'HoldCo') & (holdco_actual_result_df['account'] == account_title) & (holdco_actual_result_df['period'] == current_month_date_obj)].iloc[0]['value']/-1000.0;
        else:
            selected_worksheet.cell(row = row_number, column = 3).value = holdco_actual_result_df.loc[(holdco_actual_result_df['entity'] == 'HoldCo') & (holdco_actual_result_df['account'] == account_title) & (holdco_actual_result_df['period'] == current_month_date_obj)].iloc[0]['value']/1000.0;

        selected_worksheet.cell(row = row_number, column = 4).value = holdco_budget_result_df.loc[(holdco_budget_result_df['entity'] == 'HoldCo') & (holdco_budget_result_df['account'] == account_title) & (holdco_budget_result_df['period'] == current_month_date_obj)].iloc[0]['value']/1000.0;

        if account_title != 'Hedge P&L':
            selected_worksheet.cell(row = row_number, column = 6).value = sum(list(holdco_actual_result_df.loc[(holdco_actual_result_df['entity'] == 'HoldCo') & (holdco_actual_result_df['account'] == account_title) & (holdco_actual_result_df['period'] <= current_month_date_obj)]['value']))/-1000.0;
        else:
            selected_worksheet.cell(row = row_number, column = 6).value = sum(list(holdco_actual_result_df.loc[(holdco_actual_result_df['entity'] == 'HoldCo') & (holdco_actual_result_df['account'] == account_title) & (holdco_actual_result_df['period'] <= current_month_date_obj)]['value']))/1000.0;

        selected_worksheet.cell(row = row_number, column = 7).value = sum(list(holdco_budget_result_df.loc[(holdco_budget_result_df['entity'] == 'HoldCo') & (holdco_budget_result_df['account'] == account_title) & (holdco_budget_result_df['period'] <= current_month_date_obj)]['value']))/1000.0;







def fill_in_cells_with_financials_month(new_file_path, actest_result_df, budget_result_df, ytd_date):
    global YTD_OPTION;
    YTD_OPTION = 'MTD';
    fill_in_cells_with_financials_ytd(new_file_path, actest_result_df, budget_result_df, ytd_date, holdco_page = False);
    YTD_OPTION = 'YTD';

def draw_var_ytd(month_header, actest_result_df, budget_result_df, selected_worksheet, option):
    selected_worksheet.sheet_view.showGridLines = False;
    if option == 'summary':
        actest_result_gas_df = actest_result_df.loc[actest_result_df['entity'].isin(GAS_PLANTS_LIST)];
        budget_result_gas_df = budget_result_df.loc[budget_result_df['entity'].isin(GAS_PLANTS_LIST)];

        actest_result_gavin_df = actest_result_df.loc[actest_result_df['entity'] == 'Gavin'];
        budget_result_gavin_df = budget_result_df.loc[budget_result_df['entity'] == 'Gavin'];

        actest_result_holdco_df = actest_result_df.loc[actest_result_df['entity'] == 'HoldCo'];
        budget_result_holdco_df = budget_result_df.loc[budget_result_df['entity'] == 'HoldCo'];

        #############################################################
        """ YTD Actuals Part  """
        #############################################################
        gas_plants_value_dict = {"column_header":'Gas Plants'};

        actest_result_gas_account_df = actest_result_gas_df.groupby(['account']);
        for account in actest_result_gas_account_df.groups:
            temp_value = sum(list(actest_result_gas_account_df.get_group(account)['value']));

            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];


            gas_plants_value_dict[account] = temp_value;



        gavin_value_dict = {"column_header":'Gavin'};

        actest_result_gavin_account_df = actest_result_gavin_df.groupby(['account']);
        for account in actest_result_gavin_account_df.groups:
            temp_value = sum(list(actest_result_gavin_account_df.get_group(account)['value']));
            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];

            gavin_value_dict[account] = temp_value;

        holdco_value_dict = {"column_header":'HoldCo'};

        actest_result_holdco_account_df = actest_result_holdco_df.groupby(['account']);
        for account in actest_result_holdco_account_df.groups:
            temp_value = sum(list(actest_result_holdco_account_df.get_group(account)['value']));
            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];

            holdco_value_dict[account] = temp_value;


        column_header_total_dict = {'Gavin':'Lightstone'}
        total_value_dict = {"column_header":column_header_total_dict['Gavin']};

        actest_result_total_account_df = actest_result_df.groupby(['account']);
        for account in actest_result_total_account_df.groups:
            temp_value = sum(list(actest_result_total_account_df.get_group(account)['value']));
            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];

            total_value_dict[account] = temp_value;




        start_col = 1;
        start_row = 1;


        gas_plants_col = start_col + 2;
        for row_number in range(2,34):
            if row_number == 2:
                selected_worksheet.cell(row = row_number, column = gas_plants_col).value = month_header + " " + YTD_OPTION;
                continue;
            if row_number == 3:
                selected_worksheet.cell(row = row_number, column = gas_plants_col).value = gas_plants_value_dict['column_header'];
                selected_worksheet.cell(row = row_number, column = gas_plants_col+1).value = gavin_value_dict['column_header'];
                selected_worksheet.cell(row = row_number, column = gas_plants_col+2).value = holdco_value_dict['column_header'];
                selected_worksheet.cell(row = row_number, column = gas_plants_col+3).value = total_value_dict['column_header'];
                continue;
            current_account = selected_worksheet.cell(row = row_number, column = 1).value;
            selected_worksheet.cell(row = row_number, column = gas_plants_col).value = gas_plants_value_dict[current_account]/1000.0 if current_account in gas_plants_value_dict else 0.0;
            selected_worksheet.cell(row = row_number, column = gas_plants_col+1).value = gavin_value_dict[current_account]/1000.0 if current_account in gavin_value_dict else 0.0;
            selected_worksheet.cell(row = row_number, column = gas_plants_col+2).value =holdco_value_dict[current_account]/1000.0 if current_account in holdco_value_dict else 0.0;
            selected_worksheet.cell(row = row_number, column = gas_plants_col+3).value = total_value_dict[current_account]/1000.0 if current_account in total_value_dict else 0.0;
        #############################################################
        #############################################################




        #############################################################
        """ YTD Budget result  """
        #############################################################
        gas_plants_value_dict = {"column_header":'Gas Plants'};

        budget_result_gas_account_df = budget_result_gas_df.groupby(['account']);
        for account in budget_result_gas_account_df.groups:
            temp_value = sum(list(budget_result_gas_account_df.get_group(account)['value']));
            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];

            gas_plants_value_dict[account] = temp_value;



        gavin_value_dict = {"column_header":'Gavin'};

        budget_result_gavin_account_df = budget_result_gavin_df.groupby(['account']);
        for account in budget_result_gavin_account_df.groups:
            temp_value = sum(list(budget_result_gavin_account_df.get_group(account)['value']));

            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];

            gavin_value_dict[account] = temp_value;


        holdco_value_dict = {"column_header":'HoldCo'};

        budget_result_holdco_account_df = budget_result_holdco_df.groupby(['account']);
        for account in budget_result_holdco_account_df.groups:
            temp_value = sum(list(budget_result_holdco_account_df.get_group(account)['value']));

            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];

            holdco_value_dict[account] = temp_value;


        total_value_dict = {"column_header":column_header_total_dict['Gavin']};

        budget_result_total_account_df = budget_result_df.groupby(['account']);
        for account in budget_result_total_account_df.groups:
            temp_value = sum(list(budget_result_total_account_df.get_group(account)['value']));

            if account in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account];

            total_value_dict[account] = temp_value;

        budget_col = start_col + 6;
        for row_number in range(2,34):
            if row_number == 2:
                selected_worksheet.cell(row = row_number, column = budget_col).value = "Budget " + YTD_OPTION;
                continue;
            if row_number == 3:
                selected_worksheet.cell(row = row_number, column = budget_col).value = gas_plants_value_dict['column_header'];
                selected_worksheet.cell(row = row_number, column = budget_col+1).value = gavin_value_dict['column_header'];
                selected_worksheet.cell(row = row_number, column = budget_col+2).value = holdco_value_dict['column_header'];
                selected_worksheet.cell(row = row_number, column = budget_col+3).value = total_value_dict['column_header'];
                continue;
            current_account = selected_worksheet.cell(row = row_number, column = 1).value;
            selected_worksheet.cell(row = row_number, column = budget_col).value = gas_plants_value_dict[current_account]/1000.0 if current_account in gas_plants_value_dict else 0.0;
            selected_worksheet.cell(row = row_number, column = budget_col+1).value = gavin_value_dict[current_account]/1000.0 if current_account in gavin_value_dict else 0.0;
            selected_worksheet.cell(row = row_number, column = budget_col+2).value =holdco_value_dict[current_account]/1000.0 if current_account in holdco_value_dict else 0.0;
            selected_worksheet.cell(row = row_number, column = budget_col+3).value = total_value_dict[current_account]/1000.0 if current_account in total_value_dict else 0.0;
        #############################################################
        #############################################################



        #############################################################
        """ YTD Variance result  """
        #############################################################

        variance_col = start_col + 10;

        selected_worksheet.cell(row = 2, column = variance_col).value = "Variance " + YTD_OPTION;

        row_number = 3;

        selected_worksheet.cell(row = row_number, column = variance_col).value = gas_plants_value_dict['column_header'];
        selected_worksheet.cell(row = row_number, column = variance_col+1).value = gavin_value_dict['column_header'];
        selected_worksheet.cell(row = row_number, column = variance_col+2).value = holdco_value_dict['column_header'];
        selected_worksheet.cell(row = row_number, column = variance_col+3).value = total_value_dict['column_header'];


        #############################################################
        #############################################################

    if option == 'gas_plants':
        actest_result_gas_df = actest_result_df.loc[actest_result_df['entity'].isin(GAS_PLANTS_LIST)];
        budget_result_gas_df = budget_result_df.loc[budget_result_df['entity'].isin(GAS_PLANTS_LIST)];

        column_header_plus = '';
        if option == 'YTD':
            column_header_plus = ' ' + YTD_OPTION;

        grouped_actest_df = actest_result_gas_df.groupby(['entity']);

        start_col_actual = 3;
        start_col_budget = 7;
        start_col_variance = 11;
        for entity in GAS_PLANTS_LIST:
            account_value_dict = {};
            budget_value_dict = {};
            entity_actest_df = grouped_actest_df.get_group(entity);
            account_entity_actest_df = entity_actest_df.groupby(['account']);

            for account in account_entity_actest_df.groups:
                account_actual_df = account_entity_actest_df.get_group(account);
                temp_value = sum(list(account_actual_df['value']));
                if account in VALUE_SIGN_DICT:
                    temp_value = temp_value * VALUE_SIGN_DICT[account];

                account_value_dict[account] = temp_value;


                temp_value = sum(list(budget_result_df.loc[(budget_result_df['entity']==entity) & (budget_result_df['account']==account)]['value']));
                if account in VALUE_SIGN_DICT:
                    temp_value = temp_value * VALUE_SIGN_DICT[account];

                budget_value_dict[account] = temp_value;



            for row_number in range(3,34):
                if row_number == 3:
                    selected_worksheet.cell(row = row_number, column = start_col_actual).value = entity;
                    selected_worksheet.cell(row = row_number, column = start_col_budget).value = entity;
                    selected_worksheet.cell(row = row_number, column = start_col_variance).value = entity;

                    continue;
                account_in_template = selected_worksheet.cell(row = row_number, column = 1).value;
                selected_worksheet.cell(row = row_number, column = start_col_actual).value = account_value_dict[account_in_template]/1000.0 if account_in_template in account_value_dict else 0.0;
                selected_worksheet.cell(row = row_number, column = start_col_budget).value = budget_value_dict[account_in_template]/1000.0 if account_in_template in budget_value_dict else 0.0;

            start_col_actual += 1;
            start_col_budget += 1;
            start_col_variance += 1;



        header_row_number = 2;
        col_header_number_list = [3,7,11];

        selected_worksheet.cell(row = header_row_number, column = col_header_number_list[0]).value = 'Actual' + column_header_plus;
        selected_worksheet.cell(row = header_row_number, column = col_header_number_list[1]).value = 'Budget' + column_header_plus;
        selected_worksheet.cell(row = header_row_number, column = col_header_number_list[2]).value = 'Variance' + column_header_plus;

        for row_number in range(3,34):
            if row_number == 3:
                selected_worksheet.cell(row = row_number, column = start_col_actual).value = 'Consolidated';
                selected_worksheet.cell(row = row_number, column = start_col_budget).value = 'Consolidated';
                selected_worksheet.cell(row = row_number, column = start_col_variance).value = 'Consolidated';
                continue;

            account_title = selected_worksheet.cell(row = row_number, column = 1).value;

            temp_value = sum(list(actest_result_gas_df.loc[actest_result_gas_df['account'] == account_title]['value']));
            if account_title in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account_title];


            selected_worksheet.cell(row = row_number, column = start_col_actual).value = temp_value/1000.0 if account_title in list(actest_result_gas_df['account']) else 0.0;

            temp_value = sum(list(budget_result_gas_df.loc[budget_result_gas_df['account'] == account_title]['value']));
            if account_title in VALUE_SIGN_DICT:
                temp_value = temp_value * VALUE_SIGN_DICT[account_title];

            selected_worksheet.cell(row = row_number, column = start_col_budget).value = temp_value/1000.0 if account_title in list(budget_result_gas_df['account']) else 0.0;





def draw_var_per_entity(month_header, actest_result_df, budget_result_df, selected_worksheet, entity):
    selected_worksheet.sheet_view.showGridLines = False;

    actest_value_dict = {};

    account_actest_result_df = actest_result_df.groupby(['account']);
    for account in account_actest_result_df.groups:
        temp_value = sum(list(account_actest_result_df.get_group(account)['value']));
        if account in VALUE_SIGN_DICT:
            temp_value = temp_value * VALUE_SIGN_DICT[account];

        actest_value_dict[account] = temp_value;

    budget_value_dict = {};

    account_budget_result_df = budget_result_df.groupby(['account']);
    for account in account_budget_result_df.groups:
        temp_value = sum(list(account_budget_result_df.get_group(account)['value']));
        if account in VALUE_SIGN_DICT:
            temp_value = temp_value * VALUE_SIGN_DICT[account];

        budget_value_dict[account] = temp_value;




    for row_number in range(2, 34):
        if row_number == 2:
            selected_worksheet.cell(row = row_number, column = 3).value = entity + " " + month_header + " " + YTD_OPTION;
            continue;
        if row_number == 3:
            selected_worksheet.cell(row = row_number, column = 3).value = "Actual";
            selected_worksheet.cell(row = row_number, column = 4).value = 'Budget';
            selected_worksheet.cell(row = row_number, column = 5).value = 'Variance';
            continue;
        account_title = selected_worksheet.cell(row = row_number, column = 1).value;
        selected_worksheet.cell(row = row_number, column = 3).value = actest_value_dict[account_title]/1000.0 if account_title in actest_value_dict else 0.0;
        selected_worksheet.cell(row = row_number, column = 4).value = budget_value_dict[account_title]/1000.0 if account_title in budget_value_dict else 0.0;
