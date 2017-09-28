from openpyxl import *;
from openpyxl.utils import get_column_letter;
from openpyxl.writer.write_only import WriteOnlyCell;
from openpyxl.comments import Comment;
from openpyxl.styles import *;

import sys;
import os;
import ast;
import shutil;
import datetime;
from pathlib import Path;


sys.path.insert(0, str(Path(__file__).parents[1])+r'/utility');
import date_utils;

sys.path.insert(0, str(Path(__file__).parents[1])+r'/database');
import db_controller;




def letters_of_credit_report(conn_ins, company_name):
    """ Letters of Credit """
    ############################################################################################
    ############################################################################################


    print ("=================   generating letters of credit report...");

    report_date = str(datetime.datetime.now().date());


    lc_result_df, letters_of_credit_df = db_controller.get_lc_facilities_letters_of_credit(conn_ins, company_name);

    # output_report_file_path = "output_report/" + company_name + "_letters_of_credit.xlsx";
    output_report_file_path = str(Path(__file__).parents[2]) + r'/reports/' + company_name + "_letters_of_credit " + report_date + ".xlsx";
    create_letter_of_credit_report(output_report_file_path, lc_result_df, letters_of_credit_df, company_name);


    ############################################################################################
    ############################################################################################


def create_letter_of_credit_report(output_report_file_path, lc_result_df, letters_of_credit_df, company_name):

    lc_workbook = Workbook();
    lc_worksheet = lc_workbook.get_active_sheet();
    lc_worksheet.title = 'Letter Of Credit';

    tlc_result_list = [];
    revl_result_list = [];
    instrument_id_tlc = company_name + " TLC";
    instrument_id_rev = company_name + " Revolver";
    tlc_lc_df = lc_result_df.loc[lc_result_df['instrument_id'] == instrument_id_tlc];
    revl_lc_df = lc_result_df.loc[lc_result_df['instrument_id'] == instrument_id_rev];

    for i in range(0, len(tlc_lc_df)):
        letters_of_credit_issued = sum(letters_of_credit_df.loc[(letters_of_credit_df['issuer']==tlc_lc_df.iloc[i]['issuer']) & (letters_of_credit_df['instrument_id']==tlc_lc_df.iloc[i]['instrument_id']) ]['amount']);
        purpose_list = list(letters_of_credit_df.loc[(letters_of_credit_df['issuer']==tlc_lc_df.iloc[i]['issuer']) & (letters_of_credit_df['instrument_id']==tlc_lc_df.iloc[i]['instrument_id']) ]['description']);
        sublimit_lc = tlc_lc_df.iloc[i]['amount'];
        available_capacity = sublimit_lc - letters_of_credit_issued;
        issuer = tlc_lc_df.iloc[i]['issuer'];
        tlc_result_list.append([issuer, sublimit_lc, letters_of_credit_issued, available_capacity, purpose_list])

    for i in range(0, len(revl_lc_df)):
        letters_of_credit_issued = sum(letters_of_credit_df.loc[(letters_of_credit_df['issuer']==revl_lc_df.iloc[i]['issuer']) & (letters_of_credit_df['instrument_id']==revl_lc_df.iloc[i]['instrument_id']) ]['amount']);
        purpose_list = list(letters_of_credit_df.loc[(letters_of_credit_df['issuer']==revl_lc_df.iloc[i]['issuer']) & (letters_of_credit_df['instrument_id']==revl_lc_df.iloc[i]['instrument_id']) ]['description']);
        sublimit_lc = revl_lc_df.iloc[i]['amount'];
        available_capacity = sublimit_lc - letters_of_credit_issued;
        issuer = revl_lc_df.iloc[i]['issuer'];
        revl_result_list.append([issuer, sublimit_lc, letters_of_credit_issued, available_capacity, purpose_list])



    tlc_result_list.append(["Total Facility", sum(list(zip(*tlc_result_list))[1]), sum(list(zip(*tlc_result_list))[2]), sum(list(zip(*tlc_result_list))[3]),'']);
    revl_result_list.append(["Total Facility", sum(list(zip(*revl_result_list))[1]), sum(list(zip(*revl_result_list))[2]), sum(list(zip(*revl_result_list))[3]),'']);



    sumup_border = Border(left=Side(border_style=None,color='FF000000'),
                    right=Side(border_style=None,color='FF000000'),
                    top=Side(border_style="thin",color='FF000000'),
                    bottom=Side(border_style="double",color='FF000000'),
                    diagonal=Side(border_style=None,color='FF000000'),
                    diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                    vertical=Side(border_style=None,color='FF000000'),
                    horizontal=Side(border_style=None,color='FF000000'));



    start_row = 7;
    start_col = 2;

    tlc_start_row = start_row;
    tlc_header_row = tlc_start_row - 2;
    tlc_start_col = start_col;
    for item in tlc_result_list:
        for item_i in item[:-1]:
            lc_worksheet.cell(row=tlc_start_row, column=tlc_start_col).value = item_i;
            if item.index(item_i) != 0:
                if tlc_result_list.index(item) != 0 and tlc_result_list.index(item) < len(tlc_result_list) - 1:
                    lc_worksheet.cell(row=tlc_start_row, column=tlc_start_col).number_format = '_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)';
                else:
                    lc_worksheet.cell(row=tlc_start_row, column=tlc_start_col).number_format = '_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)';
                    if tlc_result_list.index(item) == len(tlc_result_list) - 1:
                        lc_worksheet.cell(row=tlc_start_row, column=tlc_start_col).border = sumup_border;
            tlc_start_col += 1;
            lc_worksheet.cell(row=tlc_start_row, column=tlc_start_col).value = '';
            tlc_start_col += 1;
        flag = 0;
        for last_item_i in item[-1]:
            flag += 1;
            lc_worksheet.cell(row=tlc_start_row, column=tlc_start_col).value = last_item_i;
            tlc_start_row += 1;


        if flag > 0:
            tlc_start_row -= 1;

        tlc_start_row += 1;
        tlc_start_col = start_col;




    revl_start_row = tlc_start_row + 5;
    revl_header_row = revl_start_row - 2;
    revl_start_col = start_col;

    for item in revl_result_list:

        for item_i in item[:-1]:
            lc_worksheet.cell(row=revl_start_row, column=revl_start_col).value = item_i;
            if item.index(item_i) != 0:
                if revl_result_list.index(item) != 0 and revl_result_list.index(item) < len(revl_result_list) - 1:
                    lc_worksheet.cell(row=revl_start_row, column=revl_start_col).number_format = '_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)';
                else:
                    lc_worksheet.cell(row=revl_start_row, column=revl_start_col).number_format = '_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)';
                    if revl_result_list.index(item) == len(revl_result_list) - 1:
                        lc_worksheet.cell(row=revl_start_row, column=revl_start_col).border = sumup_border;
            revl_start_col += 1;
            lc_worksheet.cell(row=revl_start_row, column=revl_start_col).value = '';
            revl_start_col += 1;
        flag = 0;
        for last_item_i in item[-1]:
            flag += 1;
            lc_worksheet.cell(row=revl_start_row, column=revl_start_col).value = last_item_i;
            revl_start_row += 1;
        if flag > 0:
            tlc_start_row -= 1;

        revl_start_row += 1;
        revl_start_col = start_col;

    lc_worksheet.column_dimensions["B"].width = 40;
    lc_worksheet.column_dimensions["C"].width = 5;
    lc_worksheet.column_dimensions["D"].width = 25;
    lc_worksheet.column_dimensions["E"].width = 12;
    lc_worksheet.column_dimensions["F"].width = 22;
    lc_worksheet.column_dimensions["G"].width = 12;
    lc_worksheet.column_dimensions["H"].width = 20;
    lc_worksheet.column_dimensions["I"].width = 12;
    lc_worksheet.column_dimensions["J"].width = 80;





    header_font = Font(name='Calibri',
                    size=11,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000');

    scenario_font = Font(name='Calibri',
                    size=20,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000');

    scenario_border = Border(left=Side(border_style=None,color='FF000000'),
                right=Side(border_style=None,color='FF000000'),
                top=Side(border_style=None,color='FF000000'),
                bottom=Side(border_style="thick",color='FF000000'),
                diagonal=Side(border_style=None,color='FF000000'),
                diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                vertical=Side(border_style=None,color='FF000000'),
                horizontal=Side(border_style=None,color='FF000000'));


    header_alignment = Alignment(horizontal='center',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0);

    header_border = Border(left=Side(border_style=None,color='FF000000'),
                    right=Side(border_style=None,color='FF000000'),
                    top=Side(border_style=None,color='FF000000'),
                    bottom=Side(border_style="thin",color='FF000000'),
                    diagonal=Side(border_style=None,color='FF000000'),
                    diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                    vertical=Side(border_style=None,color='FF000000'),
                    horizontal=Side(border_style=None,color='FF000000'));


    tlc_header_row_first_list = ['','', 'Term Letter of Credit', '', '', ''];
    tlc_header_row_second_list = ['Term Loan L/C Facility','', 'Sublimit','', 'Letters of Credit Issued','', 'Available Capacity','', 'Purpose'];

    tlc_header_start_col = start_col;
    for item in tlc_header_row_first_list:
        lc_worksheet.cell(row = tlc_header_row, column = tlc_header_start_col).value = item;
        lc_worksheet.cell(row = tlc_header_row, column = tlc_header_start_col).font = header_font;
        lc_worksheet.cell(row = tlc_header_row, column = tlc_header_start_col).alignment = header_alignment;
        tlc_header_start_col += 1;

    tlc_header_row += 1;
    tlc_header_start_col = start_col;
    for item in tlc_header_row_second_list:
        lc_worksheet.cell(row = tlc_header_row, column = tlc_header_start_col).value = item;
        lc_worksheet.cell(row = tlc_header_row, column = tlc_header_start_col).font = header_font;
        lc_worksheet.cell(row = tlc_header_row, column = tlc_header_start_col).alignment = header_alignment;
        lc_worksheet.cell(row = tlc_header_row, column = tlc_header_start_col).border = header_border;
        tlc_header_start_col += 1;



    revl_header_row_first_list = ['','', 'Revolving Letter of Credit', '', '', ''];
    revl_header_row_second_list = ['Term Loan L/C Facility','', 'Credit Sublimit','', 'Letters of Credit Issued','', 'Available Capacity','', 'Purpose'];

    revl_header_start_col = start_col;
    for item in revl_header_row_first_list:
        lc_worksheet.cell(row = revl_header_row, column = revl_header_start_col).value = item;
        lc_worksheet.cell(row = revl_header_row, column = revl_header_start_col).font = header_font;
        lc_worksheet.cell(row = revl_header_row, column = revl_header_start_col).alignment = header_alignment;
        revl_header_start_col += 1;

    revl_header_row += 1;
    revl_header_start_col = start_col;
    for item in revl_header_row_second_list:
        lc_worksheet.cell(row = revl_header_row, column = revl_header_start_col).value = item;
        lc_worksheet.cell(row = revl_header_row, column = revl_header_start_col).font = header_font;
        lc_worksheet.cell(row = revl_header_row, column = revl_header_start_col).alignment = header_alignment;
        lc_worksheet.cell(row = revl_header_row, column = revl_header_start_col).border = header_border;
        revl_header_start_col += 1;


    scenario_cell = lc_worksheet["D1"];
    scenario_cell.value = "Letters Of Credit";
    style_range(lc_worksheet, "D1:I2", font = scenario_font, alignment = header_alignment);

    for i in range(2, 11):
        lc_worksheet.cell(row = 2, column = i).border = scenario_border;

    lc_worksheet.sheet_view.showGridLines = False;

    lc_workbook.save(output_report_file_path);


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
