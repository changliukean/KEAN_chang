import openpyxl;
from openpyxl import *;
from openpyxl.utils import get_column_letter;
from openpyxl.styles import *;


import sys;
import os;
import ast;
import shutil;
import datetime;
from pathlib import Path;


sys.path.insert(0, str(Path(__file__).parents[1])+r'/utility');
import date_utils;

sys.path.insert(0, str(Path(__file__).parents[1])+r'/database');
import db_controller;

VALUE_SIGN_DICT = {'Energy Revenue':1,'Delivered Fuel Expense':-1,'Net Emissions Expense':-1,'Variable O&M Expense':-1,'Fixed Fuel':-1,'Hedge P&L':1,'Capacity Revenue':1,'Ancillary Services Revenue':1,'Misc Income':1,\
                    'Labor Expenses':-1,'Maintenance':-1,'Operations':-1,'Removal Costs':-1,'Fuel Handling':-1, 'Property Tax':-1, 'Insurance':-1, 'General & Administrative':-1,\
                    'Maintenance Capex':-1, 'Environmental Capex':-1, 'LTSA Capex':-1, 'Growth Capex':-1, 'Fixed Non-Labor Expense':-1, 'Total Fixed Costs':-1, 'Total Capex':-1}



def difftest_report(conn_ins, company, first_scenario, second_scenario, first_year, second_year,  first_version = '', second_version = '',start_month=1, end_month=12):
    """ difftest Report """
    ############################################################################################
    ############################################################################################
    print ("=================   generating difference test report...");



    start_month_str = str(start_month) if len(str(start_month)) > 1 else '0' + str(start_month);
    end_month_str = str(end_month) if len(str(end_month)) > 1 else '0' + str(end_month);

    report_date = str(datetime.datetime.now().date());



    first_report_start_date = str(first_year) + "-" + start_month_str + "-01";
    first_report_end_date = str(first_year) + "-" + end_month_str + "-31";

    second_report_start_date = str(second_year) + "-" + start_month_str + "-01";
    second_report_end_date = str(second_year) + "-" + end_month_str + "-31";


    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/diff_test_report_template.xlsx';

    print ( company , first_scenario , first_version, second_scenario, second_version, report_date );

    new_file_path = str(Path(__file__).parents[2]) + r"/reports/" + company + " diff " + first_scenario + " " + first_version + "-" + second_scenario + " " + second_version + " " + report_date + ".xlsx";

    first_result_df = db_controller.get_results_from_financials(conn_ins, company, first_scenario, data_start_date=first_report_start_date, data_end_date = first_report_end_date, version = first_version);
    second_result_df = db_controller.get_results_from_financials(conn_ins, company, second_scenario, data_start_date = second_report_start_date, data_end_date = second_report_end_date, version = second_version);
    print ("length for first results", first_scenario, first_version if first_version != '' else 'blank-version', ": ", len(first_result_df));
    print ("length for second results", second_scenario, second_version if second_version != '' else 'blank-version', ": ", len(second_result_df));

    if len(first_result_df) == 0 or len(second_result_df) == 0:
        print ("One of the target result set is null, comparison stops.");
        return 0;

    # sys.exit();

    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';

    create_copy_of_template(ori_file_path, new_file_path);

    first_scenario_version = first_scenario + '_' + first_version;
    second_scenario_version = second_scenario + '_' + second_version;

    fill_in_cells_with_financials_difftest(new_file_path, first_result_df, first_scenario_version, second_result_df, second_scenario_version);

    return new_file_path;

    #############################################################################################
    #############################################################################################


def create_copy_of_template(ori_file_path, new_file_path):
    shutil.copy(ori_file_path, new_file_path);



def fill_in_cells_with_financials_difftest(new_file_path, first_result_df, first_scenario_version, second_result_df, second_scenario_version):
    estimate_fill = PatternFill(fill_type='solid', start_color='FFE8E9EA', end_color='FFE8E9EA');
    wb = load_workbook(filename = new_file_path);
    wb, first_tab_name_list = fill_in_cells_with_financials(wb, first_result_df, first_scenario_version,'A');
    wb, second_tab_name_list = fill_in_cells_with_financials(wb, second_result_df, second_scenario_version, 'B');

    first_summary_worksheet = wb.copy_worksheet(wb.active)
    second_summary_worksheet = wb.copy_worksheet(wb.active);
    get_summary(first_summary_worksheet, first_tab_name_list, first_scenario_version, 'A');
    get_summary(second_summary_worksheet, second_tab_name_list, second_scenario_version, 'B');

    wb = get_diff(wb, first_tab_name_list, 'A', first_scenario_version, second_tab_name_list, 'B', second_scenario_version);

    number_of_diff_tabs = len(first_tab_name_list);

    wb._sheets = [wb._sheets[0]] + [wb._sheets[i] for i in range(-number_of_diff_tabs,0)] +  wb._sheets[1:-number_of_diff_tabs];

    wb.save(new_file_path);

def fill_in_cells_with_financials(workbook, financials_result_df, scenario, tab_diff):

    wb = workbook;
    ws = wb.active;
    grouped_by_entity_df = financials_result_df.groupby(['entity']);
    entity_name_list = list(grouped_by_entity_df.groups.keys());

    temp_entity_name_list = [];
    for item in entity_name_list:
        if item != 'HoldCo':
            temp_entity_name_list.append(item + tab_diff);

    temp_entity_name_list.append("HoldCo" + tab_diff);


    entity_name_list = temp_entity_name_list; """ reorder the list, make sure holdco is the last tab """



    for selected_entity in entity_name_list:
        print (selected_entity);
        # selected_worksheet = wb.create_sheet(title=selected_entity);
        selected_worksheet = wb.copy_worksheet(ws);
        selected_worksheet.title = selected_entity;

        img = openpyxl.drawing.image.Image(LOGO_PATH)
        img.anchor(selected_worksheet['B1']);
        selected_worksheet.add_image(img);


        selected_worksheet.cell(row = 1, column = 13).value = selected_entity;
        selected_worksheet.cell(row = 1, column = 6).value = scenario;


        try:
            selected_financials_result_df = grouped_by_entity_df.get_group(selected_entity[:-1]);
        except:
            selected_worksheet.sheet_view.showGridLines = False;
            break;

        print (len(selected_financials_result_df));



        start_col = 'A';
        start_row = 4;
        end_row = 33;

        month_list = [];
        month_list_flag = 0;
        for row_index in range(start_row, end_row+1):
            account = ws[start_col+str(row_index)].value;
            account_value_df = selected_financials_result_df.loc[selected_financials_result_df['account'] == account];
            account_value_df = account_value_df.sort_values(['period']);
            row_value_list = [];
            # print (selected_worksheet.title, "    ",account, "    " , len(account_value_df));
            if month_list_flag == 0 and len(account_value_df) > 0:
                month_list = list(account_value_df['period']);
                # print (selected_entity);
                month_list_flag = 1;

            for i in range(0, len(account_value_df)):
                # print (account_value_df.iloc[i]['period'], account_value_df.iloc[i]['value']);
                row_value_list.append(account_value_df.iloc[i]['value']/1000.0);


            for col_index in range(3,3+len(row_value_list)):
                list_index = col_index-3;
                selected_worksheet.cell(row=row_index, column = col_index).value = row_value_list[list_index];

        for col_index in range(3, 3+len(month_list)):
            list_index = col_index - 3;
            selected_worksheet.cell(row = 3, column=col_index).value = date_utils.calc_forecast_monthly_headers(month_list[list_index].year, month_list[list_index].month);




        selected_worksheet.sheet_view.showGridLines = False;


    # temp_blank = wb.copy_worksheet(ws);
    # temp_blank.title = 'current active';
    # first_worksheet = get_summary(ws, entity_name_list, scenario);



    # wb.save(new_file_path);
    return wb, entity_name_list;

def get_summary(first_worksheet, entity_name_list, scenario, tab_diff):
    sum_string = "=SUM(";
    sum_string = sum_string + entity_name_list[0] + ":" + entity_name_list[-1] + "!"

    img = openpyxl.drawing.image.Image(LOGO_PATH)
    img.anchor(first_worksheet['B1']);
    first_worksheet.add_image(img);

    first_worksheet.title = 'Summary'+tab_diff;

    first_worksheet.cell(row = 1, column = 6).value = scenario;
    first_worksheet.cell(row = 1, column = 13).value = 'Summary';

    for row_index in range(4,34):
        for col_index in range(3,15):
            first_worksheet.cell(row=row_index, column=col_index).value = sum_string + get_column_letter(col_index) + str(row_index) + ")";

    cell_string = "=" + entity_name_list[0] + "!";
    for col_index in range(3,15):
        first_worksheet.cell(row=2, column=col_index).value = cell_string + get_column_letter(col_index) + str(2) ;

    for col_index in range(3,15):
        first_worksheet.cell(row=3, column=col_index).value = cell_string + get_column_letter(col_index) + str(3) ;
    first_worksheet.sheet_view.showGridLines = False;

    return first_worksheet;

def get_diff(workbook, first_tab_name_list, first_tab_diff, first_scenario, second_tab_name_list, second_tab_diff, second_scenario):


    for i in range(0, len(first_tab_name_list)):
        first_worksheet = workbook.copy_worksheet(workbook.active);
        img = openpyxl.drawing.image.Image(LOGO_PATH)
        img.anchor(first_worksheet['B1']);
        first_worksheet.add_image(img);

        first_worksheet.title = first_tab_name_list[i][:-1] + first_tab_diff + " diff " + second_tab_diff;

        print (first_tab_name_list[i][:-1]);
        second_tab_name_item = [item for item in second_tab_name_list if first_tab_name_list[i][:-1] in item][0];

        if not second_tab_name_item:
            break;


        first_worksheet.cell(row = 1, column = 6).value = "Diff " + first_tab_name_list[i][:-1] + " " + first_scenario + " - " + second_tab_name_item[:-1] + " " + second_scenario ;
        # first_worksheet.cell(row = 1, column = 13).value = 'Summary';

        for row_index in range(4,34):
            for col_index in range(3,15):
                first_worksheet.cell(row=row_index, column=col_index).value = "=" + first_tab_name_list[i] + "!" + get_column_letter(col_index) + str(row_index) + "-" + second_tab_name_item + "!" + get_column_letter(col_index) + str(row_index);

        cell_string = "=" + first_tab_name_list[0] + "!";
        for col_index in range(3,15):
            first_worksheet.cell(row=2, column=col_index).value = cell_string + get_column_letter(col_index) + str(2) ;

        for col_index in range(3,15):
            first_worksheet.cell(row=3, column=col_index).value = cell_string + get_column_letter(col_index) + str(3) ;
        first_worksheet.sheet_view.showGridLines = False;



    first_worksheet = workbook.active;
    img = openpyxl.drawing.image.Image(LOGO_PATH)
    img.anchor(first_worksheet['B1']);
    first_worksheet.add_image(img);

    first_worksheet.title = "Summary" + first_tab_diff + " diff " + second_tab_diff;

    first_worksheet.cell(row = 1, column = 6).value = "Diff " + "Summary " + first_scenario + " - " + second_scenario ;
    # first_worksheet.cell(row = 1, column = 13).value = 'Summary';

    for row_index in range(4,34):
        for col_index in range(3,15):
            first_worksheet.cell(row=row_index, column=col_index).value = "=" + "Summary" + first_tab_diff + "!" + get_column_letter(col_index) + str(row_index) + "-" + "Summary" + second_tab_diff + "!" + get_column_letter(col_index) + str(row_index);

    cell_string = "=" + "Summary" + first_tab_diff + "!";
    for col_index in range(3,15):
        first_worksheet.cell(row=2, column=col_index).value = cell_string + get_column_letter(col_index) + str(2) ;

    for col_index in range(3,15):
        first_worksheet.cell(row=3, column=col_index).value = cell_string + get_column_letter(col_index) + str(3) ;
    first_worksheet.sheet_view.showGridLines = False;


    return workbook;
