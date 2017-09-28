import xlsxwriter;
from openpyxl import *;
from openpyxl.utils import get_column_letter;
from openpyxl.writer.write_only import WriteOnlyCell;
from openpyxl.comments import Comment;
from openpyxl.styles import *;

import sys;
import os;
import ast;
import shutil;
import datetime;
from pathlib import Path;


sys.path.insert(0, str(Path(__file__).parents[1])+r'/utility');
import date_utils;

sys.path.insert(0, str(Path(__file__).parents[1])+r'/database');
import db_controller;


COMPANY_ENTITY_ORDER_DICT = {
                                'Lightstone':{'Gavin':0,'Lawrenceburg':1,'Waterford':2,'Darby':3}

                            }


border = Border(left=Side(border_style="thin",color='FF000000'),right=Side(border_style="thin",color='FF000000'),
                top=Side(border_style="thin",color='FF000000'),bottom=Side(border_style="thin",color='FF000000'),
                vertical=Side(border_style="thin",color='FF000000'),horizontal=Side(border_style="thin",color='FF000000'));



def support_pxq_report(conn_ins, company, amr_scenario, pxq_pnl_ready_list, pxq_mtd_ready_list, pxq_ytd_ready_list, pxq_forecast_ready_list, forecast_monthly_list, budget_monthly_list):
    """ P times Q support report """
    ############################################################################################
    ############################################################################################

    print ("=================   generating p times q support report...");

    report_date = str(datetime.datetime.now().date());

    output_report_directory = str(Path(__file__).parents[2]) + r'/reports/support/' + company + '/' + amr_scenario + '/';

    if not os.path.exists(output_report_directory):
        os.makedirs(output_report_directory);

    output_report_file_path = output_report_directory + company + "_" + amr_scenario + "_pqanalysis_" + report_date + ".xlsx";
    # print (output_report_file_path);
    create_pqanalysis_report(output_report_file_path, company, amr_scenario, pxq_pnl_ready_list, pxq_mtd_ready_list, pxq_ytd_ready_list, pxq_forecast_ready_list, forecast_monthly_list, budget_monthly_list);

    return output_report_file_path;


    ############################################################################################
    ############################################################################################

def create_pqanalysis_report(output_report_file_path, company, amr_scenario, pxq_pnl_ready_list, pxq_mtd_ready_list, pxq_ytd_ready_list, pxq_forecast_ready_list, af_monthly_list, budget_monthly_list):

    pxq_workbook = Workbook();

    pxq_workbook = pxq_pnl_sheet(pxq_workbook, pxq_pnl_ready_list);

    pxq_workbook = pxq_mtd_sheet(pxq_workbook, company, pxq_mtd_ready_list);

    pxq_workbook = pxq_ytd_sheet(pxq_workbook, company, pxq_ytd_ready_list);

    pxq_workbook = pxq_forecast_sheet(pxq_workbook, company, pxq_forecast_ready_list);

    pxq_workbook = pxq_forecast_budget_monthly_sheets(pxq_workbook, company, amr_scenario, af_monthly_list, budget_monthly_list);

    pxq_workbook.remove(pxq_workbook.active);

    pxq_workbook.save(output_report_file_path);


def pxq_forecast_budget_monthly_sheets(pxq_workbook, company, amr_scenario, af_monthly_list, budget_monthly_list):
    pxq_af_monthly_worksheet = pxq_workbook.copy_worksheet(pxq_workbook.active);
    pxq_af_monthly_worksheet.title = 'A+F Monthly Data';
    pxq_budget_monthly_worksheet = pxq_workbook.copy_worksheet(pxq_workbook.active);
    pxq_budget_monthly_worksheet.title = 'Budget Monthly Data';



    start_row_number = 6;

    current_year = amr_scenario.split(" ")[0][-2:];

    header_list = ['Entity','Account Item', 'Jan-'+current_year, 'Feb-'+current_year, 'Mar-'+current_year, 'Apr-'+current_year, 'May-'+current_year, 'Jun-'+current_year, 'Jul-'+current_year, 'Aug-'+current_year, 'Sep-'+current_year, 'Oct-'+current_year, 'Nov-'+current_year, 'Dec-'+current_year];

    for i in range(0, len(header_list)):
        pxq_af_monthly_worksheet.cell(row = start_row_number, column = i+1).value = header_list[i];

    start_row_number += 1;

    for temp_af_monthly_list in af_monthly_list:
        for i in range(0,len(temp_af_monthly_list)):
            for j in range(0, len(temp_af_monthly_list[i])):
                pxq_af_monthly_worksheet.cell(row = start_row_number, column = j+1).value = temp_af_monthly_list[i][j];
            start_row_number += 1;




    start_row_number = 6;

    for i in range(0, len(header_list)):
        pxq_budget_monthly_worksheet.cell(row = start_row_number, column = i+1).value = header_list[i];

    start_row_number += 1;

    for temp_budget_monthly_list in budget_monthly_list:
        for i in range(0,len(temp_budget_monthly_list)):
            for j in range(0, len(temp_budget_monthly_list[i])):
                pxq_budget_monthly_worksheet.cell(row = start_row_number, column = j+1).value = temp_budget_monthly_list[i][j];
            start_row_number += 1;




    return pxq_workbook;





def pxq_forecast_sheet(pxq_workbook, company, pxq_forecast_ready_list):
    print ("--------------- forecast sheet------------------");

    print (len(pxq_forecast_ready_list));

    pxq_forecast_worksheet = pxq_workbook.copy_worksheet(pxq_workbook.active);
    pnl_tab_title = 'PQ Anaylysis - P&L';
    forecast_tab_title = 'PQ Analysis - Estimate';
    pxq_forecast_worksheet.title = forecast_tab_title;

    pxq_forecast_worksheet['A2'] = 'As of: 2017-08-31';
    pxq_forecast_worksheet['A4'] = 'AMR: Price * Quantity Analysis';

    start_row_number = 6;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'ESTIMATE';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Realized Prices';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'Energy Revenue';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 8).value = 'Total';

    start_row_number += 1;

    forecast_ready_list = pxq_forecast_ready_list[0];

    for i in range(0, len(forecast_ready_list)):
        for j in range(0, len(forecast_ready_list[i])):
            pxq_forecast_worksheet.cell(row = start_row_number, column = j+1).value = forecast_ready_list[i][j];
        start_row_number += 1;



    start_row_number += 4;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'BUDGET';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Realized Prices';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'Energy Revenue';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 8).value = 'Total';

    start_row_number += 1;

    budget_ready_list = pxq_forecast_ready_list[1];

    for i in range(0, len(budget_ready_list)):
        for j in range(0, len(budget_ready_list[i])):
            pxq_forecast_worksheet.cell(row = start_row_number, column = j+1).value = budget_ready_list[i][j];
        start_row_number += 1;


    ##################

    start_row_number += 4;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'VARIANCE';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Realized Prices';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'Energy Revenue';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 8).value = 'Total';

    start_row_number += 1;

    variance_ready_list = pxq_forecast_ready_list[2];

    for i in range(0, len(variance_ready_list)):
        for j in range(0, len(variance_ready_list[i])):
            pxq_forecast_worksheet.cell(row = start_row_number, column = j+1).value = variance_ready_list[i][j];
        start_row_number += 1;


    ###############################

    start_row_number += 4;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'Energy Revenue: PxQ Analysis';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Realized Prices';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'Energy Revenue';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 8).value = 'Total';

    start_row_number += 1;

    pxq_ready_list = pxq_forecast_ready_list[3];

    for i in range(0, len(pxq_ready_list)):
        for j in range(0, len(pxq_ready_list[i])):
            pxq_forecast_worksheet.cell(row = start_row_number, column = j+1).value = pxq_ready_list[i][j];
        start_row_number += 1;


    #################################

    start_row_number += 4;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'Delivered Fuel Expense: PxQ Analysis';

    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'Actual';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'Budget';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 6).value = 'Delta';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 8).value = 'PxQ';


    start_row_number += 1;

    pxq_forecast_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 2).value = 'Fuel Burn';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 3).value = 'Fuel Price';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 4).value = 'Fuel Burn';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 5).value = 'Fuel Price';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 6).value = 'Fuel Burn';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 7).value = 'Fuel Price';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 8).value = 'Price';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 9).value = 'Quantity';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 10).value = 'Total';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 11).value = 'Actual Delievered Fuel Expense';
    pxq_forecast_worksheet.cell(row = start_row_number, column = 12).value = 'Budget Delievered Fuel Expense';
    start_row_number += 1;

    delivered_fuel_ready_list = pxq_forecast_ready_list[4];

    for i in range(0, len(delivered_fuel_ready_list)):
        for j in range(0, len(delivered_fuel_ready_list[i])):
            pxq_forecast_worksheet.cell(row = start_row_number, column = j+1).value = delivered_fuel_ready_list[i][j];
        start_row_number += 1;


    pxq_forecast_worksheet.sheet_view.showGridLines = False;
    return pxq_workbook;

def pxq_ytd_sheet(pxq_workbook, company, pxq_ytd_ready_list):
    print ("--------------- ytd sheet------------------");

    print (len(pxq_ytd_ready_list));

    pxq_ytd_worksheet = pxq_workbook.copy_worksheet(pxq_workbook.active);
    pnl_tab_title = 'PQ Anaylysis - P&L';
    ytd_tab_title = 'PQ Analysis - YTD';
    pxq_ytd_worksheet.title = ytd_tab_title;

    pxq_ytd_worksheet['A2'] = 'As of: 2017-07-31';
    pxq_ytd_worksheet['A4'] = 'AMR: Price * Quantity Analysis';


    start_row_number = 6;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'ACTUALS';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 5).value = 'BUDGET';

    start_row_number += 1;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 3).value = 'Realized Prices';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 4).value = 'Energy Revenue';

    pxq_ytd_worksheet.cell(row = start_row_number, column = 5).value = 'Generation';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 8).value = 'Realized Prices';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 11).value = 'Energy Revenue';

    start_row_number += 1;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Total';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 3).value = 'Total';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 4).value = 'Total';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 5).value = 'On Peak';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 6).value = 'Off Peak';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 7).value = 'Total';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 8).value = 'On Peak';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 9).value = 'Off Peak';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 10).value = 'ATC';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 11).value = 'On Peak';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 12).value = 'Off Peak';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 13).value = 'Total';


    start_row_number += 1;

    act_budget_ready_list = pxq_ytd_ready_list[0];

    for i in range(0, len(act_budget_ready_list)):
        for j in range(0, len(act_budget_ready_list[i])):
            pxq_ytd_worksheet.cell(row = start_row_number, column = j+1).value = act_budget_ready_list[i][j];
        start_row_number += 1;


    start_row_number += 4;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'VARIANCE';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 5).value = 'Energy Revenue: PxQ Analysis';

    start_row_number += 1;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 3).value = 'Realized Prices';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 4).value = 'Energy Revenue';

    pxq_ytd_worksheet.cell(row = start_row_number, column = 5).value = 'Price';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 6).value = 'Quantity';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 7).value = 'Total';

    start_row_number += 1;

    var_pxq_ready_list = pxq_ytd_ready_list[1];

    for i in range(0, len(var_pxq_ready_list)):
        for j in range(0, len(var_pxq_ready_list[i])):
            pxq_ytd_worksheet.cell(row = start_row_number, column = j+1).value = var_pxq_ready_list[i][j];
        start_row_number += 1;


    start_row_number += 4;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Delievered Fuel Expense: PxQ Analysis';

    start_row_number += 1;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Actual';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 4).value = 'Budget';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 6).value = 'Delta';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 8).value = 'PxQ';

    start_row_number += 1;
    pxq_ytd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Fuel Burn';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 3).value = 'Fuel Price';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 4).value = 'Fuel Burn';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 5).value = 'Fuel Price';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 6).value = 'Fuel Burn';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 7).value = 'Fuel Price';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 8).value = 'Price';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 9).value = 'Quantity';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 10).value = 'Total';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 11).value = 'Actual Delivered Fuel Expense';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 12).value = 'Budget Delivered Fuel Expense';

    start_row_number += 1;

    del_fuel_pxq_ready_list = pxq_ytd_ready_list[2];

    for i in range(0, len(del_fuel_pxq_ready_list)):
        for j in range(0, len(del_fuel_pxq_ready_list[i])):
            pxq_ytd_worksheet.cell(row = start_row_number, column = j+1).value = del_fuel_pxq_ready_list[i][j];
        start_row_number += 1;





    start_row_number += 4;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Capacity Factor Build-up';

    start_row_number += 1;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Capacity';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 3).value = 'Period Hours';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 4).value = 'Generation';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 5).value = 'Capacity Factor';


    start_row_number += 1;

    capacity_factor_pxq_ready_list = pxq_ytd_ready_list[3];

    for i in range(0, len(capacity_factor_pxq_ready_list)):
        for j in range(0, len(capacity_factor_pxq_ready_list[i])):
            pxq_ytd_worksheet.cell(row = start_row_number, column = j+1).value = capacity_factor_pxq_ready_list[i][j];
        start_row_number += 1;


    start_row_number += 4;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'EAF Build-up';

    start_row_number += 1;

    pxq_ytd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 2).value = 'Service Hours';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 3).value = 'Reserve Shutdown Hours';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 4).value = 'Available Hours';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 5).value = 'Equivalent Plant Derate Hours';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 6).value = 'Equivalent Unit Derate Hours';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 7).value = 'Period Hours';
    pxq_ytd_worksheet.cell(row = start_row_number, column = 8).value = 'EAF';

    start_row_number += 1;

    eaf_pxq_ready_list = pxq_ytd_ready_list[4];

    for i in range(0, len(eaf_pxq_ready_list)):
        for j in range(0, len(eaf_pxq_ready_list[i])):
            pxq_ytd_worksheet.cell(row = start_row_number, column = j+1).value = eaf_pxq_ready_list[i][j];
        start_row_number += 1;


    pxq_ytd_worksheet.sheet_view.showGridLines = False;

    return pxq_workbook;

def pxq_mtd_sheet(pxq_workbook, company, pxq_mtd_ready_list):
    print ("--------------- mtd sheet------------------");

    print (len(pxq_mtd_ready_list));

    # for item in pxq_mtd_ready_list:
    #     print (item);

    pxq_mtd_worksheet = pxq_workbook.copy_worksheet(pxq_workbook.active);
    pnl_tab_title = 'PQ Anaylysis - P&L';
    mtd_tab_title = 'PQ Analysis - MTD';
    pxq_mtd_worksheet.title = mtd_tab_title;

    start_row_number = 6;
    pxq_mtd_worksheet['A4'] = 'AMR: Price * Quantity Analysis';

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'ACTUALS';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Realized Prices';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Energy Revenue';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = 'Total';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 9).value = 'ATC Generation';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 10).value = 'ATC Realized Price';

    actuals_list = [item for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] in ['Total Generation']];
    actuals_list.sort(key=lambda val:(COMPANY_ENTITY_ORDER_DICT[company][val[1]]));

    # for item in actuals_list:
    #     print (item);

    start_row_number += 1;
    row_sections_list = [];

    row_sections_list.append(start_row_number);


    for i in range(0, len(actuals_list)):
        current_entity = actuals_list[i][1];
        current_value = actuals_list[i][4];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = current_entity;

        cell_value_str = "=(\'" + pnl_tab_title + "\'!B" + str(start_row_number) + "/SUM(\'" + pnl_tab_title + "\'!$B" + str(start_row_number) + ":$C" + str(start_row_number) + "))*\'" + mtd_tab_title + "\'!$I" + str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = cell_value_str;

        cell_value_str = "=(\'" + pnl_tab_title + "\'!C" + str(start_row_number) + "/SUM(\'" + pnl_tab_title + "\'!$B" + str(start_row_number) + ":$C" + str(start_row_number) + "))*\'" + mtd_tab_title + "\'!$I" + str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = cell_value_str;

        cell_value_str = "=(\'" + pnl_tab_title + "\'!D" + str(start_row_number) + "/\'" + pnl_tab_title + "\'!$I" + str(start_row_number) + ")*\'" + mtd_tab_title + "\'!$J" + str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = cell_value_str;

        cell_value_str = "=(\'" + pnl_tab_title + "\'!E" + str(start_row_number) + "/\'" + pnl_tab_title + "\'!$I" + str(start_row_number) + ")*\'" + mtd_tab_title + "\'!$J" + str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = cell_value_str;

        cell_value_str = "=B" + str(start_row_number) + "*D" + str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = cell_value_str;

        cell_value_str = "=C" + str(start_row_number) + "*E" + str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = cell_value_str;


        pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = [item[4] for item in pxq_mtd_ready_list if item[0] == 'ACTUALS' and item[1] == current_entity and item[2] == 'Energy Revenue'][0]/1000.0;

        pxq_mtd_worksheet.cell(row = start_row_number, column = 9).value = current_value;

        cell_value_str = "=H" + str(start_row_number) + "/I" + str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 10).value = cell_value_str;



        start_row_number += 1;

    row_sections_list.append(start_row_number-1);

    for col in [1,2,3,6,7,8,9]:
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "Total";
            continue;
        col_letter = get_column_letter(col)
        pos_str = col_letter + str(start_row_number);
        pxq_mtd_worksheet[pos_str] = '=SUM(' + col_letter + str(row_sections_list[0]) + ":" + col_letter + str(row_sections_list[1]) + ")";

    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = "=F" + str(start_row_number) + "/" + "B" + str(start_row_number);
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = "=G" + str(start_row_number) + "/" + "C" + str(start_row_number);
    pxq_mtd_worksheet.cell(row = start_row_number, column = 10).value = "=H" + str(start_row_number) + "/" + "I" + str(start_row_number);


    start_row_number += 4;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'BUDGET';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Realized Prices';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Energy Revenue';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = 'Total';
    start_row_number += 1;


    temp_budget_list = [item for item in pxq_mtd_ready_list if item[0] == 'BUDGET' and item[2] == 'Fossil Fuel Consumed'];

    temp_budget_list.sort(key=lambda val:(COMPANY_ENTITY_ORDER_DICT[company][val[1]]));

    budget_list =  [item for item in pxq_mtd_ready_list if item[0] == 'BUDGET'];
    row_sections_list.append(start_row_number);

    for i in range(0, len(temp_budget_list)):
        current_entity = temp_budget_list[i][1];

        pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = current_entity;

        on_peak_gen = [item[4] for item in budget_list if item[0] == 'BUDGET' and item[2] == 'On-Peak Generation' and item[1] == current_entity][0];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = on_peak_gen;

        off_peak_gen = [item[4] for item in budget_list if item[0] == 'BUDGET' and item[2] == 'Off-Peak Generation' and item[1] == current_entity][0];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = off_peak_gen;

        on_peak_price = [item[4] for item in budget_list if item[0] == 'BUDGET' and item[2] == 'Realized On-Peak Power Price' and item[1] == current_entity][0];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = on_peak_price;


        off_peak_price = [item[4] for item in budget_list if item[0] == 'BUDGET' and item[2] == 'Realized Off-Peak Power Price' and item[1] == current_entity][0];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = off_peak_price;


        pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = "=B"+str(start_row_number)+"*D"+str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = "=C"+str(start_row_number)+"*E"+str(start_row_number);
        pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = "=SUM(F"+str(start_row_number)+":G"+str(start_row_number) + ")";

        start_row_number += 1;

    row_sections_list.append(start_row_number-1);

    for col in [1,2,3,6,7,8]:
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "Total";
            continue;
        col_letter = get_column_letter(col)
        pos_str = col_letter + str(start_row_number);
        pxq_mtd_worksheet[pos_str] = '=SUM(' + col_letter + str(row_sections_list[2]) + ":" + col_letter + str(row_sections_list[3]) + ")";

    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = "=F" + str(start_row_number) + "/" + "B" + str(start_row_number);
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = "=G" + str(start_row_number) + "/" + "C" + str(start_row_number);


    start_row_number += 4;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'VARIANCE';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Realized Prices';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Energy Revenue';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = 'Total';
    start_row_number += 1;

    # print (row_sections_list);

    row_sections_list.append(start_row_number);

    for i in range(0, len(temp_budget_list)):
        for j in range(1,9):
            if j == 1:
                pxq_mtd_worksheet.cell(row = start_row_number, column = j).value = temp_budget_list[i][j];
            else:
                pxq_mtd_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j)) + str(list(range(row_sections_list[0], row_sections_list[1]+1))[i]) + "-" + str(get_column_letter(j)) + str(list(range(row_sections_list[2], row_sections_list[3]+1))[i])

        start_row_number += 1;

    row_sections_list.append(start_row_number-1);

    for col in [1,2,3,4,5,6,7,8]:
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "Total";
            continue;
        col_letter = get_column_letter(col)
        pos_str = col_letter + str(start_row_number);
        pxq_mtd_worksheet[pos_str] = '=' + col_letter + str(row_sections_list[1]+1) + "-" + col_letter + str(row_sections_list[3]+1) ;



    start_row_number += 4;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'P x Q Analysis ';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Price';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Quantity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Total';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Price';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = 'Quantity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = 'Total';
    start_row_number += 1;

    row_sections_list.append(start_row_number);

    for i in range(0, len(temp_budget_list)):
        for j in range(1,9):
            if j == 1:
                pxq_mtd_worksheet.cell(row = start_row_number, column = j).value = temp_budget_list[i][j];
            else:
                if j in [2,3]:
                    pxq_mtd_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j+2)) + str(list(range(row_sections_list[4], row_sections_list[5]+1))[i]) + "*" + str(get_column_letter(j)) + str(list(range(row_sections_list[2], row_sections_list[3]+1))[i]);
                if j in [4,5]:
                    pxq_mtd_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j-2)) + str(list(range(row_sections_list[4], row_sections_list[5]+1))[i]) + "*" + str(get_column_letter(j)) + str(list(range(row_sections_list[0], row_sections_list[1]+1))[i]);
                if j in [6]:
                    pxq_mtd_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(2)) + str(start_row_number) + "+" + str(get_column_letter(3)) + str(start_row_number);
                if j in [7]:
                    pxq_mtd_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(4)) + str(start_row_number) + "+" + str(get_column_letter(5)) + str(start_row_number);
                if j in [8]:
                    pxq_mtd_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j-2)) + str(start_row_number) + "+" + str(get_column_letter(j-1)) + str(start_row_number);

        start_row_number += 1;

    row_sections_list.append(start_row_number-1);

    for col in [1,2,3,4,5,6,7,8]:
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "Total";
            continue;
        col_letter = get_column_letter(col)
        pos_str = col_letter + str(start_row_number);
        pxq_mtd_worksheet[pos_str] = '=SUM(' + col_letter + str(row_sections_list[6]) + ":" + col_letter + str(row_sections_list[7]) + ")";


    start_row_number += 4;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Capacity Factor Build-up';

    start_row_number += 1;
    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Capacity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Period Hours';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Generation';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = 'Capacity Factor';

    start_row_number += 1;
    capacity_factor_list = [];

    for i in range(0, len(temp_budget_list)):
        current_entity = temp_budget_list[i][1];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = current_entity;
        current_capacity = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'Budget Capacity' and item[1] == current_entity][0];
        current_period_hours = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'Period Hours' and item[1] == current_entity][0];
        current_generation = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'Total Generation' and item[1] == current_entity][0];

        pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = current_capacity;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = current_period_hours;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = current_generation;

        pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = current_generation/(current_period_hours * current_capacity)*1000;
        capacity_factor_list.append([current_entity, current_generation/(current_period_hours * current_capacity)*1000]);
        start_row_number += 1;

    for col in range(1,6):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Gas Plants';
        if col in [2,4]:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "=SUM("+get_column_letter(col)+str(start_row_number-3)+":"+get_column_letter(col)+str(start_row_number-1)+")";
        if col == 3:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "="+get_column_letter(col)+str(start_row_number-1);
        if col == 5:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "="+get_column_letter(4)+str(start_row_number) + "/(" + get_column_letter(2) + str(start_row_number) + "*" + get_column_letter(3) + str(start_row_number) + ")*1000";

    start_row_number += 1;

    for col in range(1,6):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Gavin';
        if col in [2,4]:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "=SUM("+get_column_letter(col)+str(start_row_number-5)+":"+get_column_letter(col)+str(start_row_number-5)+")";
        if col == 3:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "="+get_column_letter(col)+str(start_row_number-1);
        if col == 5:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "="+get_column_letter(4)+str(start_row_number) + "/(" + get_column_letter(2) + str(start_row_number) + "*" + get_column_letter(3) + str(start_row_number) + ")*1000";

    start_row_number += 1;
    company_capacity_row_number = start_row_number;

    for col in range(1,6):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Lightstone';
        if col in [2,4]:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "=SUM("+get_column_letter(col)+str(start_row_number-2)+":"+get_column_letter(col)+str(start_row_number-1)+")";
        if col == 3:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "="+get_column_letter(col)+str(start_row_number-1);
        if col == 5:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "="+get_column_letter(4)+str(start_row_number) + "/(" + get_column_letter(2) + str(start_row_number) + "*" + get_column_letter(3) + str(start_row_number) + ")*1000";


    start_row_number += 4;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'EAF Build-up';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Service Hours';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Reserve Shutdown Hours';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Available Hours';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = 'Equivalent Plant Derate Hours';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Equivalent Unit Derate Hours';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = 'Period Hours';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = 'EAF';
    start_row_number += 1;

    eaf_list = [];

    for i in range(0, len(temp_budget_list)):
        current_entity = temp_budget_list[i][1];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = current_entity;
        current_sh = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'SH' and item[1] == current_entity][0];
        current_rs = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'RS' and item[1] == current_entity][0];
        current_epdh = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'EPDH' and item[1] == current_entity][0];
        current_eudh = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'EUDH' and item[1] == current_entity][0];
        current_period_hours = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'PH' and item[1] == current_entity][0];

        pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = current_sh;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = current_rs;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = current_rs + current_sh;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = current_epdh;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = current_eudh;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = current_period_hours;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = (current_sh + current_rs - current_epdh - current_eudh)/current_period_hours;

        eaf_list.append([current_entity,  (current_sh + current_rs - current_epdh - current_eudh)/current_period_hours]);

        start_row_number += 1;



    capacity_list = [item for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'Budget Capacity'];
    period_hours_list = [item for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'Period Hours'];

    # print ([ capacity_list[i][4] * eaf_list[i][1] for i in range(0, len(capacity_list)) if capacity_list[i][1] != 'Gavin' ]);


    for col in range(1,9):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Gas Plants';
        if col == 8:

            sum_product = sum([ capacity_list[i][4] * eaf_list[i][1] for i in range(0, len(capacity_list)) if capacity_list[i][1] != 'Gavin' ]) / sum([ capacity_list[i][4] for i in range(0, len(capacity_list)) if capacity_list[i][1] != 'Gavin' ]);
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = sum_product;

    start_row_number += 1;

    for col in range(1,9):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Gavin';
        if col == 8:
            sum_product = sum([ capacity_list[i][4] * eaf_list[i][1] for i in range(0, len(capacity_list)) if capacity_list[i][1] == 'Gavin' ]) / sum([ capacity_list[i][4] for i in range(0, len(capacity_list)) if capacity_list[i][1] == 'Gavin' ]);
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = sum_product;

    start_row_number += 1;


    for col in range(1,9):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Lightstone';
        if col == 8:
            sum_product = sum([ capacity_list[i][4] * eaf_list[i][1] for i in range(0, len(capacity_list)) ]) / sum([ capacity_list[i][4] for i in range(0, len(capacity_list))]);
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = sum_product;


    start_row_number += 4;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Available GWh';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Actual GWh';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Available & Uneconomic GWh';
    start_row_number += 1;

    for i in range(0, len(temp_budget_list)):
        current_entity = temp_budget_list[i][1];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = current_entity;
        current_avail_gwh = [item[4] for item in capacity_list if item[1] == current_entity][0] * [item[4] for item in period_hours_list if item[1] == current_entity][0] * [item[1] for item in eaf_list if item[0] == current_entity][0] / 1000;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = current_avail_gwh;
        current_actual_gwh = [item[4] for item in capacity_list if item[1] == current_entity][0] * [item[4] for item in period_hours_list if item[1] == current_entity][0] * [item[1] for item in capacity_factor_list if item[0] == current_entity][0] / 1000;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = current_actual_gwh;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = current_avail_gwh - current_actual_gwh;
        start_row_number += 1;

    for col in range(1, 5):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Total';
        else:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "=SUM(" + get_column_letter(col) + str(start_row_number-len(temp_budget_list)) + ":" + get_column_letter(col) + str(start_row_number-1) + ")";

    start_row_number += 1;

    for col in range(1, 5):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Hours';
        else:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "=" + get_column_letter(col) + str(start_row_number-1) + "*1000/" + get_column_letter(col) + str(company_capacity_row_number) ;



    start_row_number += 4;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Budget Capacity Factor';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Budget Availablity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Budget GWh (Avail & Econ.)';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = 'Budget GWh Avail';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Budget (Avail & Unecon.)';

    start_row_number += 1;

    # for item in pxq_mtd_ready_list:
    #     print (item);

    for i in range(0, len(temp_budget_list)):
        current_entity = temp_budget_list[i][1];
        pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = current_entity;

        current_budget_cap_fac = [item[4] for item in pxq_mtd_ready_list if item[0] == 'BUDGET' and item[1] == current_entity and item[2] == 'Total Generation'][0] / ( [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[1] == current_entity and item[2] == 'Budget Capacity'][0] * [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[1] == current_entity and item[2] == 'Period Hours'][0]) * 1000;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = current_budget_cap_fac;

        """ HERE we have a quick fix !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"""
        """ but this is very dangerous!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"""
        current_eaf = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[1] == current_entity and item[2] == 'Equivalent Availablity Factor (EAF) %'][1];

        pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = current_eaf;

        current_budget_gwh_avail_econ = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[1] == current_entity and item[2] == 'Budget Capacity'][0] * current_budget_cap_fac;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = current_budget_gwh_avail_econ;

        current_budget_gwh_avail = [item[4] for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[1] == current_entity and item[2] == 'Budget Capacity'][0] * current_eaf;
        pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = current_budget_gwh_avail;

        pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = current_budget_gwh_avail - current_budget_gwh_avail_econ;

        start_row_number += 1;

    for col in range(1,7):
        if col == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = 'Total';
        else:
            pxq_mtd_worksheet.cell(row = start_row_number, column = col).value = "=SUM(" + get_column_letter(col) + str(start_row_number - len(temp_budget_list)) + ":" + get_column_letter(col) + str(start_row_number - 1) + ")";

    start_row_number += 4;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Delivered Fuel Expense: PxQ Analysis';
    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Actual';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Budget';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Delta';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = 'PxQ';

    start_row_number += 1;

    pxq_mtd_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 2).value = 'Fuel Burn';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 3).value = 'Fuel Price';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 4).value = 'Fuel Burn';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 5).value = 'Fuel Price';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 6).value = 'Fuel Burn';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 7).value = 'Fuel Price';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 8).value = 'Price';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 9).value = 'Quantity';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 10).value = 'Total';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 11).value = 'Actual Delivered Fuel Expense';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 12).value = 'Budget Delivered Fuel Expense';
    pxq_mtd_worksheet.cell(row = start_row_number, column = 13).value = 'Delta Delivered Fuel Expense';

    start_row_number += 1;

    # for item in pxq_mtd_ready_list:
    #     print (item);

    pxq_fuel_burn_list = [item for item in pxq_mtd_ready_list if item[0] == 'PXQ' and item[2] == 'Fuel Burn'];

    pxq_fuel_burn_list.sort(key=lambda val:(COMPANY_ENTITY_ORDER_DICT[company][val[1]]));

    # for item in pxq_fuel_burn_list:
    #     print (item);

    budget_fuel_burn_list = [item for item in pxq_mtd_ready_list if item[0] == 'BUDGET' and item[2] == 'Fossil Fuel Consumed'];

    budget_fuel_burn_list.sort(key=lambda val:(COMPANY_ENTITY_ORDER_DICT[company][val[1]]));

    # for item in budget_fuel_burn_list:
    #     print (item);

    actual_del_fuel_list = [item for item in pxq_mtd_ready_list if item[0] == 'ACTUALS' and item[2] == 'Delivered Fuel Expense'];
    actual_del_fuel_list.sort(key=lambda val:(COMPANY_ENTITY_ORDER_DICT[company][val[1]]));

    budget_del_fuel_list = [item for item in pxq_mtd_ready_list if item[0] == 'BUDGET' and item[2] == 'Delivered Fuel Expense'];
    budget_del_fuel_list.sort(key=lambda val:(COMPANY_ENTITY_ORDER_DICT[company][val[1]]));


    # for item in actual_del_fuel_list:
    #     print (item);
    #
    # for item in budget_del_fuel_list:
    #     print (item);

    del_fuel_result_list = [];

    for i in range(0, len(pxq_fuel_burn_list)):
        current_entity = pxq_fuel_burn_list[i][1];
        current_act_fuel = [item[4] for item in pxq_fuel_burn_list if item[1] == current_entity][0];
        current_budget_fuel = [item[4] for item in budget_fuel_burn_list if item[1] == current_entity][0];
        current_act_del_fuel = -[item[4] for item in actual_del_fuel_list if item[1] == current_entity][0]/1000;
        current_budget_del_fuel = -[item[4] for item in budget_del_fuel_list if item[1] == current_entity][0]/1000;
        current_act_fuel_price = current_act_del_fuel / current_act_fuel;
        current_budget_fuel_price = current_budget_del_fuel / current_budget_fuel;
        current_fuel_delta = current_act_fuel - current_budget_fuel;
        current_price_delta = current_act_fuel_price - current_budget_fuel_price;
        current_pxq_price = current_act_fuel_price * current_fuel_delta;
        current_pxq_quantity = current_budget_fuel * current_price_delta;
        current_pxq_total = current_pxq_price + current_pxq_quantity;
        current_del_fuel_delta = current_act_del_fuel - current_budget_del_fuel;

        row_list = [current_entity,current_act_fuel,current_act_fuel_price,current_budget_fuel,current_budget_fuel_price,
                    current_fuel_delta, current_price_delta, current_pxq_price, current_pxq_quantity,current_pxq_total,
                    current_act_del_fuel,current_budget_del_fuel, current_del_fuel_delta];

        del_fuel_result_list.append(row_list);
        for j in range(0, len(row_list)):
            pxq_mtd_worksheet.cell(row = start_row_number, column = j+1).value = row_list[j];
        start_row_number += 1;

    for i in range(1,14):
        if i == 1:
            pxq_mtd_worksheet.cell(row = start_row_number, column = i).value = 'Total';
        if i in [2,4,6,8,9,10,11,12,13]:
            pxq_mtd_worksheet.cell(row = start_row_number, column = i).value = sum(list(zip(*del_fuel_result_list))[i-1]);
        if i == 3:
            pxq_mtd_worksheet.cell(row = start_row_number, column = i).value = sum(list(zip(*del_fuel_result_list))[10]) / sum(list(zip(*del_fuel_result_list))[1]);
        if i == 5:
            pxq_mtd_worksheet.cell(row = start_row_number, column = i).value = sum(list(zip(*del_fuel_result_list))[11]) / sum(list(zip(*del_fuel_result_list))[3]);
        if i == 7:
            pxq_mtd_worksheet.cell(row = start_row_number, column = i).value = sum(list(zip(*del_fuel_result_list))[10]) / sum(list(zip(*del_fuel_result_list))[1]) - sum(list(zip(*del_fuel_result_list))[11]) / sum(list(zip(*del_fuel_result_list))[3]);


    pxq_mtd_worksheet.sheet_view.showGridLines = False;

    return pxq_workbook;

def pxq_pnl_sheet(pxq_workbook, pxq_pnl_ready_list):
    print ("--------------- pnl sheet------------------");
    pxq_pnl_worksheet = pxq_workbook.copy_worksheet(pxq_workbook.active);
    pxq_pnl_worksheet.title = 'PQ Anaylysis - P&L';
    #
    # for item in pxq_pnl_ready_list:
    #     print (item);


    start_row_number = 6;
    pxq_pnl_worksheet['A4'] = 'P&L: Price * Quantity Analysis';

    budget_list = [item for item in pxq_pnl_ready_list if item[0] == 'Budget'];
    actuals_list = [item for item in pxq_pnl_ready_list if item[0] == 'PI Actuals'];

    pxq_pnl_worksheet.cell(row = start_row_number, column = 1).value = 'ACTUALS';
    start_row_number += 1;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = 'Realized Prices';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 6).value = 'Energy Revenue';
    start_row_number += 1;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 8).value = 'Total';

    pxq_pnl_worksheet.cell(row = start_row_number, column = 9).value = 'ATC Realized Price';

    start_row_number += 1;

    sum_start_row = 0;
    sum_end_row = 0;

    current_col = 1;

    for i in range(0, len(actuals_list)):
        if i==0:
            sum_start_row = start_row_number;
        if i == len(actuals_list)-1:
            sum_end_row = start_row_number;

        for j in range(1, len(actuals_list[i])):
            pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = actuals_list[i][j];
            current_col = j;

        current_col += 1;
        pxq_pnl_worksheet.cell(row = start_row_number, column = current_col).value = "=B" + str(start_row_number) + "*D" + str(start_row_number);
        current_col += 1;
        pxq_pnl_worksheet.cell(row = start_row_number, column = current_col).value = "=C" + str(start_row_number) + "*E" + str(start_row_number);
        current_col += 1;
        pxq_pnl_worksheet.cell(row = start_row_number, column = current_col).value = "=F" + str(start_row_number) + "+G" + str(start_row_number);
        current_col += 1;
        pxq_pnl_worksheet.cell(row = start_row_number, column = current_col).value = "=H" + str(start_row_number) + "/SUM(B" + str(start_row_number) + ":C" + str(start_row_number) + ")";


        start_row_number += 1;

    for col in [1,2,3,6,7,8,9]:
        if col == 1:
            pxq_pnl_worksheet.cell(row = start_row_number, column = col).value = "Total";
            continue;
        col_letter = get_column_letter(col)
        pos_str = col_letter + str(start_row_number);
        pxq_pnl_worksheet[pos_str] = '=SUM(' + col_letter + str(sum_start_row) + ":" + col_letter + str(sum_end_row) + ")";
        if col == 9:
            pxq_pnl_worksheet.cell(row = start_row_number, column = current_col).value = "=H" + str(start_row_number) + "/SUM(B" + str(start_row_number) + ":C" + str(start_row_number) + ")";


    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = "=F" + str(start_row_number) + "/" + "B" + str(start_row_number);
    pxq_pnl_worksheet.cell(row = start_row_number, column = 5).value = "=G" + str(start_row_number) + "/" + "C" + str(start_row_number);





    start_row_number += 4;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 1).value = 'BUDGET';
    start_row_number += 1;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = 'Realized Prices';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 6).value = 'Energy Revenue';
    start_row_number += 1;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 8).value = 'Total';
    start_row_number += 1;

    row_sections_list = [sum_start_row,sum_end_row];

    sum_start_row = 0;
    sum_end_row = 0;

    for i in range(0, len(budget_list)):
        if i == 0:
            sum_start_row = start_row_number;
        if i == len(budget_list) - 1:
            sum_end_row = start_row_number;

        current_col = 0;
        for j in range(1, len(budget_list[i])):
            pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = budget_list[i][j];
            current_col = j;

        current_col += 1;
        pxq_pnl_worksheet.cell(row = start_row_number, column = current_col).value = "=B" + str(start_row_number) + "*D" + str(start_row_number);
        current_col += 1;
        pxq_pnl_worksheet.cell(row = start_row_number, column = current_col).value = "=C" + str(start_row_number) + "*E" + str(start_row_number);
        current_col += 1;
        pxq_pnl_worksheet.cell(row = start_row_number, column = current_col).value = "=F" + str(start_row_number) + "+G" + str(start_row_number);

        start_row_number += 1;

    for col in [1,2,3,6,7,8]:
        if col == 1:
            pxq_pnl_worksheet.cell(row = start_row_number, column = col).value = "Total";
            continue;
        col_letter = get_column_letter(col)
        pos_str = col_letter + str(start_row_number);
        pxq_pnl_worksheet[pos_str] = '=SUM(' + col_letter + str(sum_start_row) + ":" + col_letter + str(sum_end_row) + ")";

    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = "=F" + str(start_row_number) + "/" + "B" + str(start_row_number);
    pxq_pnl_worksheet.cell(row = start_row_number, column = 5).value = "=G" + str(start_row_number) + "/" + "C" + str(start_row_number);


    row_sections_list.append(sum_start_row);
    row_sections_list.append(sum_end_row);

    start_row_number += 4;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 1).value = 'VARIANCE';
    start_row_number += 1;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 2).value = 'Generation';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = 'Realized Prices';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 6).value = 'Energy Revenue';
    start_row_number += 1;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 6).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 7).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 8).value = 'Total';
    start_row_number += 1;

    # print (row_sections_list);

    row_sections_list.append(start_row_number);

    for i in range(0, len(actuals_list)):
        for j in range(1,len(actuals_list[i])):
            if j == 1:
                pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = actuals_list[i][j];
            else:
                pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j)) + str(list(range(row_sections_list[0], row_sections_list[1]+1))[i]) + "-" + str(get_column_letter(j)) + str(list(range(row_sections_list[2], row_sections_list[3]+1))[i])
        for j in range(len(actuals_list[i]),len(actuals_list[i])+3):
            pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j)) + str(list(range(row_sections_list[0], row_sections_list[1]+1))[i]) + "-" + str(get_column_letter(j)) + str(list(range(row_sections_list[2], row_sections_list[3]+1))[i])

        start_row_number += 1;

    row_sections_list.append(start_row_number-1);

    for col in [1,2,3,4,5,6,7,8]:
        if col == 1:
            pxq_pnl_worksheet.cell(row = start_row_number, column = col).value = "Total";
            continue;
        col_letter = get_column_letter(col)
        pos_str = col_letter + str(start_row_number);
        pxq_pnl_worksheet[pos_str] = '=' + col_letter + str(row_sections_list[1]+1) + "-" + col_letter + str(row_sections_list[3]+1) ;





    start_row_number += 4;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 1).value = 'P x Q Analysis ';
    start_row_number += 1;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 2).value = 'Price';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = 'Quantity';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 6).value = 'Total';
    start_row_number += 1;

    pxq_pnl_worksheet.cell(row = start_row_number, column = 1).value = 'Entity';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 2).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 3).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 4).value = 'On Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 5).value = 'Off Peak';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 6).value = 'Price';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 7).value = 'Quantity';
    pxq_pnl_worksheet.cell(row = start_row_number, column = 8).value = 'Total';


    start_row_number += 1;

    row_sections_list.append(start_row_number);
    for i in range(0, len(actuals_list)):
        for j in range(1,len(actuals_list[i])):
            if j == 1:
                pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = actuals_list[i][j];
            else:
                if j in [2,3]:
                    pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j+2)) + str(list(range(row_sections_list[4], row_sections_list[5]+1))[i]) + "*" + str(get_column_letter(j)) + str(list(range(row_sections_list[2], row_sections_list[3]+1))[i]);
                if j in [4,5]:
                    pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j-2)) + str(list(range(row_sections_list[4], row_sections_list[5]+1))[i]) + "*" + str(get_column_letter(j)) + str(list(range(row_sections_list[0], row_sections_list[1]+1))[i]);

        for j in range(len(actuals_list[i]),len(actuals_list[i])+3):
            if j in [6]:
                pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(2)) + str(start_row_number) + "+" + str(get_column_letter(3)) + str(start_row_number);
            if j in [7]:
                pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(4)) + str(start_row_number) + "+" + str(get_column_letter(5)) + str(start_row_number);
            if j in [8]:
                pxq_pnl_worksheet.cell(row = start_row_number, column = j).value = "=" + str(get_column_letter(j-2)) + str(start_row_number) + "+" + str(get_column_letter(j-1)) + str(start_row_number);

        start_row_number += 1;

    row_sections_list.append(start_row_number-1);

    for col in [1,2,3,4,5,6,7,8]:
        if col == 1:
            pxq_pnl_worksheet.cell(row = start_row_number, column = col).value = "Total";
            continue;
        col_letter = get_column_letter(col)
        pos_str = col_letter + str(start_row_number);
        pxq_pnl_worksheet[pos_str] = '=SUM(' + col_letter + str(row_sections_list[6]) + ":" + col_letter + str(row_sections_list[7]) + ")";



    pxq_pnl_worksheet.cell(row = start_row_number, column=1).value = 'ATC Realized Price';




    print (row_sections_list);





    """ number_format """
    number_format_price = '"$"#,##0.00_);("$"#,##0.00)';
    number_format_nonprice = '#,##0.00_);(#,##0.00)';
    for i in range(0,len(row_sections_list)-3):
        if i == 0:
            price_cell_range_str = 'I' + str(row_sections_list[i]) + ':I' + str(row_sections_list[i+1]+1);
            style_range(pxq_pnl_worksheet,price_cell_range_str,number_format=number_format_price);

        price_cell_range_str = 'D' + str(row_sections_list[i]) + ':E' + str(row_sections_list[i+1]+1);
        style_range(pxq_pnl_worksheet,price_cell_range_str,number_format=number_format_price);
        nonprice_cell_range_str = 'B' + str(row_sections_list[i]) + ':C' + str(row_sections_list[i+1]+1);
        style_range(pxq_pnl_worksheet,nonprice_cell_range_str,number_format=number_format_nonprice);
        nonprice_cell_range_str = 'F' + str(row_sections_list[i]) + ':H' + str(row_sections_list[i+1]+1);
        style_range(pxq_pnl_worksheet,nonprice_cell_range_str,number_format=number_format_nonprice);

    for i in range(len(row_sections_list)-2,len(row_sections_list)-1):
        nonprice_cell_range_str = 'B' + str(row_sections_list[i]) + ':H' + str(row_sections_list[i+1]+1);
        style_range(pxq_pnl_worksheet,nonprice_cell_range_str,number_format=number_format_nonprice);

    """ header_format """
    center_align = Alignment(horizontal='center');
    grey_fill = PatternFill(start_color = 'FFF8F8F8');

    actual_header_cell_str = "A" + str(row_sections_list[0]-3) + ":" + "H" + str(row_sections_list[0]-3);
    merge_style_range(pxq_pnl_worksheet, actual_header_cell_str, alignment = center_align, fill = grey_fill);

    budget_header_cell_str = "A" + str(row_sections_list[2]-3) + ":" + "H" + str(row_sections_list[2]-3);
    merge_style_range(pxq_pnl_worksheet, budget_header_cell_str, alignment = center_align, fill = grey_fill);

    variance_header_cell_str = "A" + str(row_sections_list[4]-3) + ":" + "H" + str(row_sections_list[4]-3);
    merge_style_range(pxq_pnl_worksheet, variance_header_cell_str, alignment = center_align, fill = grey_fill);

    pqanalysis_header_cell_str = "A" + str(row_sections_list[6]-3) + ":" + "H" + str(row_sections_list[6]-3);
    merge_style_range(pxq_pnl_worksheet, pqanalysis_header_cell_str, alignment = center_align, fill = grey_fill);


    """ Rest of the formats to be implemented """

    """ ............  """


    for column_cells in pxq_pnl_worksheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells[6:]);
        pxq_pnl_worksheet.column_dimensions[column_cells[0].column].width = length;



    pxq_pnl_worksheet.sheet_view.showGridLines = False;
    return pxq_workbook;

def as_text(value):
    if value is None:
        return ""
    return str(value)



def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None, number_format=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
        if number_format:
            for c in row:
                c.number_format = number_format;


def merge_style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None, number_format=None):
    ws.merge_cells(cell_range);
    style_range(ws, cell_range, alignment=alignment);
