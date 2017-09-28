from openpyxl import *;
from openpyxl.utils import get_column_letter;
import openpyxl;

import sys;
import os;
import ast;
import shutil;
import pandas as pd;
import datetime;
from pathlib import Path;



sys.path.insert(0, str(Path(__file__).parents[1])+r'/utility');
import date_utils;

sys.path.insert(0, str(Path(__file__).parents[1])+r'/database');
import db_controller;



VALUE_SIGN_DICT = {'Energy Revenue':1,'Delivered Fuel Expense':-1,'Net Emissions Expense':-1,'Variable O&M Expense':-1,'Fixed Fuel':-1,'Hedge P&L':1,'Capacity Revenue':1,'Ancillary Services Revenue':1,'Misc Income':1,\
                    'Labor Expenses':-1,'Maintenance':-1,'Operations':-1,'Removal Costs':-1,'Fuel Handling':-1, 'Property Tax':-1, 'Insurance':-1, 'General & Administrative':-1,\
                    'Maintenance Capex':-1, 'Environmental Capex':-1, 'LTSA Capex':-1, 'Growth Capex':-1, 'Fixed Non-Labor Expense':-1, 'Total Fixed Costs':-1, 'Total Capex':-1}




LOGO_PATH = '';



def full_year_report(conn_ins, company,scenario, year, start_month=1, end_month=12,version = 'v5'):
    """ full year report """
    ############################################################################################
    ############################################################################################
    print ("=================   generating full year report...");
    start_month_str = str(start_month) if len(str(start_month)) > 1 else '0' + str(start_month);
    end_month_str = str(end_month) if len(str(end_month)) > 1 else '0' + str(start_month);

    report_date = str(datetime.datetime.now().date());

    report_start_date = str(year) + "-" + start_month_str + "-01";
    report_end_date = str(year) + "-" + end_month_str + "-31";
    ori_file_path = str(Path(__file__).parents[2]) + r'/templates/bvr_report_template_fy.xlsx';
    new_file_path = str(Path(__file__).parents[2]) + r'/reports/' + company + " " + scenario + " year " + str(year) + " " + report_date + ".xlsx";
    financials_result_df = db_controller.get_results_from_financials(conn_ins, company, scenario, data_start_date = report_start_date, data_end_date = report_end_date, version=version);

    print (len(financials_result_df));

    global LOGO_PATH;
    LOGO_PATH = str(Path(__file__).parents[2]) + r'/images/' + company + '_Logo.jpg';

    create_copy_of_template(ori_file_path, new_file_path);
    fy_sumup_values_dict = fill_in_cells_with_financials(new_file_path, financials_result_df, scenario);
    return new_file_path;

def create_copy_of_template(ori_file_path, new_file_path):
    shutil.copy(ori_file_path, new_file_path);



def fill_in_cells_with_financials(new_file_path, financials_result_df, scenario):


    fy_sumup_values_dict = {};
    grouped_by_entity_df = financials_result_df.groupby(['entity']);
    entity_name_list = list(grouped_by_entity_df.groups.keys());

    temp_entity_name_list = [];
    for item in entity_name_list:
        if item != 'HoldCo':
            temp_entity_name_list.append(item);

    temp_entity_name_list.append("HoldCo");


    entity_name_list = temp_entity_name_list; """ reorder the list, make sure holdco is the last tab """

    wb = load_workbook(filename = new_file_path);
    ws = wb.active;


    for selected_entity in entity_name_list:

        # selected_worksheet = wb.create_sheet(title=selected_entity);
        selected_worksheet = wb.copy_worksheet(ws);
        selected_worksheet.title = selected_entity;


        img = openpyxl.drawing.image.Image(LOGO_PATH)
        img.anchor(selected_worksheet['B1']);
        selected_worksheet.add_image(img);

        selected_worksheet.cell(row = 1, column = 13).value = selected_entity;
        selected_worksheet.cell(row = 1, column = 6).value = scenario;

        selected_financials_result_df = grouped_by_entity_df.get_group(selected_entity);
        print (len(selected_financials_result_df));



        start_col = 'A';
        start_row = 4;
        end_row = 33;

        month_list = [];
        month_list_flag = 0;
        for row_index in range(start_row, end_row+1):
            account = ws[start_col+str(row_index)].value;
            account_value_df = selected_financials_result_df.loc[selected_financials_result_df['account'] == account];
            account_value_df = account_value_df.sort_values(['period']);
            row_value_list = [];
            if month_list_flag == 0 and len(account_value_df) > 0:
                month_list = list(account_value_df['period']);
                month_list_flag = 1;

            for i in range(0, len(account_value_df)):
                # print (account_value_df.iloc[i]['period'], account_value_df.iloc[i]['value']);
                """ here we divide the exact values by 1000, cuz we dont want to mess up the values stored in db """

                temp_value = account_value_df.iloc[i]['value']/1000.0;
                if account in VALUE_SIGN_DICT:
                    temp_value =temp_value * VALUE_SIGN_DICT[account];

                row_value_list.append(temp_value);


            for col_index in range(3,3+len(row_value_list)):
                list_index = col_index-3;
                selected_worksheet.cell(row=row_index, column = col_index).value = row_value_list[list_index];

        for col_index in range(3, 3+len(month_list)):
            list_index = col_index - 3;
            selected_worksheet.cell(row = 3, column=col_index).value = date_utils.calc_forecast_monthly_headers(month_list[list_index].year, month_list[list_index].month);

            selected_worksheet.cell(row = 2, column=col_index).value = '';
            if "Budget" in scenario:
                selected_worksheet.cell(row = 2, column=col_index).value = 'Budget';
                continue;
            if month_list[list_index].year == int(scenario.split(" ")[0]):
                selected_worksheet.cell(row = 2, column=col_index).value = 'A+E';
                continue;
            if month_list[list_index].year > int(scenario.split(" ")[0]):
                selected_worksheet.cell(row = 2, column=col_index).value = 'Forecast';
                continue;

        budget_sumup_values_list = [];
        for row_index in range(4,34):
            selected_worksheet.cell(row=row_index, column = 15).value = "=SUM(C" + str(row_index) + ":N" + str(row_index) + ")";
            sumup_budget = 0.0;
            for col_index in range(3,15):
                # print (selected_entity, row_index, col_index, selected_worksheet.cell(row = row_index, column = col_index).value);
                if selected_worksheet.cell(row = row_index, column = col_index).value:
                    sumup_budget += selected_worksheet.cell(row = row_index, column = col_index).value;
                else:
                    pass;

            budget_sumup_values_list.append(sumup_budget);

        fy_sumup_values_dict[selected_entity] = budget_sumup_values_list;
        selected_worksheet.sheet_view.showGridLines = False;


    get_summary(ws, entity_name_list, scenario);
    wb.save(new_file_path);

    return fy_sumup_values_dict;


def get_summary(first_worksheet, entity_name_list, scenario):

    img = openpyxl.drawing.image.Image(LOGO_PATH)
    img.anchor(first_worksheet['B1']);
    first_worksheet.add_image(img);

    sum_string = "=SUM(";
    sum_string = sum_string + entity_name_list[0] + ":" + entity_name_list[-1] + "!"

    first_worksheet.cell(row=1, column = 6).value = scenario;
    first_worksheet.cell(row=1, column = 13).value = 'Summary';

    first_worksheet.title = 'Summary';
    for row_index in range(4,34):
        for col_index in range(3,16):
            first_worksheet.cell(row=row_index, column=col_index).value = sum_string + get_column_letter(col_index) + str(row_index) + ")";

    cell_string = "=" + entity_name_list[0] + "!";
    for col_index in range(3,15):
        first_worksheet.cell(row=2, column=col_index).value = cell_string + get_column_letter(col_index) + str(2) ;




    for col_index in range(3,16):
        first_worksheet.cell(row=3, column=col_index).value = cell_string + get_column_letter(col_index) + str(3) ;
    first_worksheet.sheet_view.showGridLines = False;
