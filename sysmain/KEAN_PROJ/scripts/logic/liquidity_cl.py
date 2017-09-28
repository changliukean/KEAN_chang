import mysql.connector
from mysql.connector import errorcode

import datetime
from datetime import datetime
from datetime import date
import calendar
from calendar import monthrange
from dateutil import relativedelta
from dateutil.relativedelta import relativedelta

import math
from math import pow

import pandas as pd

import os
import sys

path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
path_utlities = path + '/utilities/'
sys.path.insert(0, path_utlities)

import openpyxl as opx
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle


import utilities



def get_all_libor(valuation_date, instrument_id):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')

    query = "SELECT period, rate FROM libor WHERE instrument_id = %s AND valuation_date = %s "

    df = pd.read_sql(query, cnx, params=[instrument_id, valuation_date])

    return df


def get_libor(df, period):
    ''' assumes df is order by period '''
    period = date(period.year, period.month, period.day)
    curve_date = pd.to_datetime(df.iloc[0]['period'])
    if df.iloc[0]['period'] > period:
        print('Error - requested date precedes curve', period)
        return None
    elif df.iloc[-1]['period'] < period:
        print('Error - requested date exceeds curve')
        return None
    else:
        if df.iloc[0]['period'] == period:
            return df.iloc[0]['rate']
        else:
            index = 0
            while df.iloc[index]['period'] < period:
                lower_date = df.iloc[index]['period']
                lower_value = df.iloc[index]['rate']
                index += 1
            upper_date = df.iloc[index]['period']
            upper_value = df.iloc[index]['rate']
            return ((period-lower_date)/(upper_date-lower_date))*(upper_value-lower_value)+lower_value


def calc_libor_forward(df_libor_all, valuation_date, period):
    prior_period = utilities.calc_prior_month_end(period, 'date')
    period = date(period.year, period.month, period.day)
    lower_value = get_libor(df_libor_all, prior_period)
    upper_value = get_libor(df_libor_all, period)
    total_days = (period - valuation_date).days
    forward_days = monthrange(period.year, period.month)[1]
    prior_days = total_days - forward_days
    forward_rate = pow(1+upper_value/360, total_days) / pow(1+lower_value/360, prior_days)
    forward_rate = (pow(forward_rate, 1/forward_days)-1)*360

    return forward_rate


def get_all_swap(instrument_id):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')

    query = "SELECT date_start, date_end, notional, fixed_rate FROM swaps WHERE instrument_id = %s"

    df = pd.read_sql(query, cnx, params=[instrument_id])

    return df


def get_swap(df, period):
    period = date(period.year, period.month, period.day)

    if period < df.iloc[0]['date_start']:
        return None
    if period > df.iloc[-1]['date_end']:
        return None

    mask = (df['date_start'] <= period) & (df['date_end']>=period)
    notional = df[mask]['notional'].values[0]
    rate = df[mask]['fixed_rate'].values[0]

    return (notional, rate)


def build_daily_swap(valuation_date, forecast_start, forecast_end):
    swap_curve = get_all_swap('IR_Swap_1')

    df = pd.DataFrame(pd.date_range(forecast_start, forecast_end, freq='D', name='period'))
    df.set_index(['period'], inplace=True)
    df['swap_rate'] = None
    df['swap_balance'] = None
    df['calc_swap_rate'] = None
    for index, day in df.iterrows():
        #swap = get_swap(swap_curve, day[0])
        swap = get_swap(swap_curve, index)
        #swap_balance = 0 if swap is None else swap[0]
        swap_balance = 0 if swap is None else swap[0]
        swap_rate = 0 if swap is None else swap[1]
        df.set_value(index, ['swap_balance'], swap_balance)
        df.set_value(index, ['swap_rate'], swap_rate)
        df.set_value(index, ['calc_swap_rate'], swap_balance * swap_rate)

    return df


def get_inception_debt_balance(instrument_id):
    ''' assumes only using 'Actuals' scenario, entity = 'HoldCo' and instrument_id has company '''

    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    cursor = cnx.cursor(buffered=True)

    query = ("SELECT amount FROM debt WHERE instrument_id = %s")

    try:
        cursor.execute(query, (instrument_id, ))
    except mysql.connector.Error as err:
        print(err.msg)

    record = cursor.fetchone()
    if record is not None:
        return record[0]
    else:
        return None


def get_inception_debt_date(instrument_id):
    ''' assumes only using 'Actuals' scenario, entity = 'HoldCo' and instrument_id has company '''

    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    cursor = cnx.cursor(buffered=True)
    query = ("SELECT inception_date FROM debt WHERE instrument_id = %s ")

    try:
        cursor.execute(query, (instrument_id, ))
    except mysql.connector.Error as err:
        print(err.msg)

    record = cursor.fetchone()
    if record is not None:
        return record[0]
    else:
        return None


def get_amort_percent(instrument_id):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    cursor = cnx.cursor(buffered=True)
    query = ("SELECT amortization FROM debt WHERE instrument_id = %s ")

    try:
        cursor.execute(query, (instrument_id, ))
    except mysql.connector.Error as err:
        print(err.msg)

    record = cursor.fetchone()
    if record is not None:
        return record[0]
    else:
        return None


def get_debt_activity_all(instrument_id):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    query = ("SELECT payment_type, payment_date, amount FROM debt_payments WHERE instrument_id = %s")
    df = pd.read_sql(query, cnx, params=[instrument_id])

    return df


def get_debt_activity(df, activity, period):
    month_start = date(period.year, period.month, 1)
    month_end = utilities.calc_month_end(period, 'date')
    mask = (df['payment_date'] >= month_start) & (df['payment_date'] <= month_end) & (df['payment_type'] == activity)
    return df[mask]['amount'].sum()


def calc_required_amort(df, instrument_id, as_of_date):
    df_required = df.loc[:as_of_date]
    amort = 0
    for index, row in df_required.iterrows():
        if index.month % 3 == 0:
            amort += 1625000000 * 0.0025
    return amort


def get_cash():
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    cursor = cnx.cursor(buffered=True)

    baml = '%%Bank of America%%'
    db = '%%Deutsche Lightstone%%'
    db_ex = 'Deutsche Lightstone Debt Srvc Reserve Acct SC5173.4'
    query = "SELECT accounting_month, sum(ending_balance) 'Cash EOP' FROM trial_balance WHERE " \
        " (account_title LIKE %s OR account_title LIKE %s) AND account_title <> %s " \
        "GROUP BY accounting_month ORDER BY accounting_month"
    df_tb = pd.read_sql(query, cnx, params=[baml, db, db_ex])
    df_tb.set_index('accounting_month', inplace=True)
    return df_tb


def get_trial_balance(account, column, column_title):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    query = "SELECT accounting_month, sum(" + column + ") '"+ column_title + "' FROM trial_balance WHERE account_title  = %s " \
        "GROUP BY accounting_month ORDER BY accounting_month"
    df_tb = pd.read_sql(query, cnx, params=[account])
    df_tb.set_index('accounting_month', inplace=True)
    return df_tb


def calc_tax_depreciation(entity, new):
    ''' hard code rules for now - move to scenario assumptions '''
    ''' HoldCo reflects goodwill '''
    ''' new is True/False '''
    entities = {'Gavin':28, 'Waterford': 28, 'Lawrenceburg': 28, 'Darby': 20, 'HoldCo': 20}
    depreciation_rate = 1 / entities[entity]
    if new:
        return depreciation_rate / 2
    else:
        return depreciation_rate


def get_new_capex(scenario, entity, current_year):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    cursor = cnx.cursor(buffered=True)
    query = ("SELECT sum(value) FROM financials WHERE entity = %s AND scenario = %s "
            "AND account in ('Maintenance Capex', 'Environmental Capex', 'LTSA Capex', 'Growth Capex') "
            " AND period >= %s AND period <= %s")
    year_start = date(current_year, 1, 1)
    year_end = date(current_year, 12, 31)

    try:
        cursor.execute(query, (entity, scenario, year_start, year_end))
    except mysql.connector.Error as err:
        print(err.msg)

    record = cursor.fetchone()
    if record[0] is not None:
        return -record[0]
    else:
        return 0




def build_debt(df, forecast_start, forecast_end):
    ''' set initial debt balances '''
    tlb_bop = get_inception_debt_balance('Lightstone TLB')
    tlc_bop = get_inception_debt_balance('Lightstone TLC')

    df_debt_activity_tlb = get_debt_activity_all('Lightstone TLB')
    df_debt_activity_tlb['payment_date_date'] = pd.to_datetime(df_debt_activity_tlb['payment_date'])
    df_debt_activity_tlc = get_debt_activity_all('Lightstone TLC')
    df_debt_activity_tlc['payment_date_date'] = pd.to_datetime(df_debt_activity_tlc['payment_date'])
    last_activity_date_tlb = utilities.calc_month_end(df_debt_activity_tlb['payment_date_date'].max(), 'date')
    last_activity_date_tlc = utilities.calc_month_end(df_debt_activity_tlc['payment_date_date'].max(), 'date')

    for index, row in df.iterrows():
        ''' tlb first '''
        tlb_addl_borrow = 0
        tlb_prepay = 0
        tlb_amort = 0
        if date(index.year, index.month, index.day) <= max(last_activity_date_tlb, forecast_start):
            df.set_value(index, 'TLB BOP', tlb_bop)
            tlb_addl_borrow = get_debt_activity(df_debt_activity_tlb, 'additional borrowing', index)
            df.set_value(index, 'TLB Addl Borrow', tlb_addl_borrow)
            tlb_amort = get_debt_activity(df_debt_activity_tlb, 'amortization', index)
            df.set_value(index, 'TLB Amort', tlb_amort)
            tlb_prepay = get_debt_activity(df_debt_activity_tlb, 'prepayment', index)
            df.set_value(index, 'TLB Prepay', tlb_prepay)
            tlb_eop = tlb_bop + tlb_addl_borrow - tlb_amort - tlb_prepay
            df.set_value(index, 'TLB EOP', tlb_eop)
            tlb_bop = tlb_eop

        ''' tlc second '''
        if date(index.year, index.month, index.day) < max(last_activity_date_tlc, forecast_start):
            df.set_value(index, 'TLC BOP', tlc_bop)
            tlc_addl_borrow = get_debt_activity(df_debt_activity_tlc, 'additional borrowing', index)
            df.set_value(index, 'TLC Addl Borrow', tlc_addl_borrow)
            tlc_amort = get_debt_activity(df_debt_activity_tlc, 'amortization', index)
            df.set_value(index, 'TLC Amort',tlc_amort)
            tlc_prepay = get_debt_activity(df_debt_activity_tlc, 'prepayment', index)
            df.set_value(index, 'TLC Prepay',tlc_prepay)
            tlc_eop = tlc_bop + tlc_addl_borrow - tlc_amort - tlc_prepay
            df.set_value(index, 'TLC EOP', tlc_eop)
            tlc_bop = tlc_eop

        ''' forecast period '''
        if date(index.year, index.month, index.day) >= forecast_start:
            if row['TLB BOP'] <= 0:
                ''' calc amortization '''
                if index.month % 3 == 0:
                    required_amort = calc_required_amort(df, 'Lightstone TLB', index)
                    cum_prepay = df.loc[:index]['TLB Prepay'].sum()
                    cum_amort = df.loc[:index]['TLB Amort'].sum()
                    tlb_amort = max(min(1625000000*0.0025,required_amort-cum_prepay-cum_amort),0)
                    df.set_value(index, ['TLB Amort'], tlb_amort)
                df.set_value(index, 'TLB BOP', tlb_bop)
                tlb_eop = tlb_bop + tlb_addl_borrow - tlb_amort - tlb_prepay
                df.set_value(index, 'TLB EOP', tlb_bop - tlb_amort)
                tlb_bop = tlb_eop
            if row['TLC BOP'] <= 0:
                df.set_value(index, 'TLC BOP', tlc_bop)
                tlc_eop =  tlc_bop + tlc_addl_borrow - tlc_amort - tlc_prepay
                df.set_value(index, 'TLC EOP', tlc_bop)
                tlc_bop = tlc_eop

    return df


def build_libor(df_liquidity, df_libor_all, forecast_start, valuation_date):
    for index, row in df_liquidity.iterrows():
        if date(index.year, index.month, index.day) == utilities.calc_month_end(forecast_start, 'date'):
            ''' first month does not use forward curve '''
            libor = get_libor(df_libor_all, index)
            df_liquidity.set_value(index, ['LIBOR'], libor)
        elif date(index.year, index.month, index.day) > utilities.calc_month_end(forecast_start, 'date'):
            libor = calc_libor_forward(df_libor_all, valuation_date, index)
            df_liquidity.set_value(index, ['LIBOR'], libor)
    return df_liquidity


def build_swap_rate(df_liquidity, df_swap, forecast_start):
    for index, row in df_liquidity.iterrows():
        if date(index.year, index.month, index.day) >= utilities.calc_month_end(forecast_start, 'date'):
            month_start = date(index.year, index.month, 1)
            month_end = date(index.year, index.month, index.day)
            swap_balance = df_swap[month_start:month_end]['swap_balance'].sum()
            if swap_balance > 0:
                df_liquidity.set_value(index, ['Swap Avg Daily Balance'], swap_balance/index.day)
                swap_rate = df_swap[month_start:month_end]['calc_swap_rate'].sum() / swap_balance
                df_liquidity.set_value(index, ['Swap Fix Rate'], swap_rate + .045000)
            else:
                df_liquidity.set_value(index, ['Swap Avg Daily Balance'], 0)
                df_liquidity.set_value(index, ['Swap Fix Rate'], 0)
    return df_liquidity


def build_tlb_int_exp(df_liquidity, forecast_start):
    start = utilities.calc_month_end(forecast_start, 'date')

    df_liquidity.loc[start:,['TLB Int Exp']] = df_liquidity['Swap Avg Daily Balance'] * df_liquidity['Swap Fix Rate'] \
            + (df_liquidity['TLB BOP']- df_liquidity['Swap Avg Daily Balance']) * df_liquidity['TLB Float Rate']
    df_liquidity.loc[start:,['TLB Int Exp']] = df_liquidity['TLB Int Exp'] * df_liquidity.index.day / 360
    return df_liquidity


def build_tlc_int_exp(df_liquidity, forecast_start):
    start = utilities.calc_month_end(forecast_start, 'date')
    df_liquidity.loc[start:,['TLC Int Exp']] = df_liquidity['TLC BOP'] * df_liquidity['TLC Float Rate'] * df_liquidity.index.day / 360



def build_ebitda(df_liquidity, scenario, company):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    query = ("SELECT period, SUM(value) EBITDA FROM financials WHERE company = %s "
            "AND account = 'EBITDA' AND scenario = %s GROUP BY company, period ORDER BY period")
    df = pd.read_sql(query, cnx, params=[company, scenario])
    df.set_index('period', inplace = True)
    df_add = df_liquidity.add(df, fill_value=0)
    return df_add


def build_capex(df_liquidity, scenario, company):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                    user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    query = ("SELECT period, -SUM(value) Capex FROM financials WHERE company = %s "
            "AND scenario = %s AND account IN ('Maintenance Capex', 'Environmental Capex', 'LTSA Capex', 'Growth Capex') "
            "GROUP BY company, period ORDER BY period")
    df = pd.read_sql(query, cnx, params=[company, scenario])
    df.set_index('period', inplace = True)
    df_add = df_liquidity.add(df, fill_value=0)
    return df_add


def build_cash_actuals(df_liquidity, df_cash):
    ''' must set initial cash balance manually for now '''
    cash_bop = 56900000
    for index, row in df_cash.iterrows():
        df_liquidity.set_value(index, 'Cash BOP', cash_bop)
        cash_eop = row['Cash EOP']
        df_liquidity.set_value(index, 'Cash EOP', cash_eop)
        cash_bop = cash_eop
    return df_liquidity


def build_dsra(df_liquidity):
    df_dsra = get_trial_balance('Deutsche Lightstone Debt Srvc Reserve Acct SC5173.4', 'ending_balance', 'DSRA EOP')
    for index, row in df_dsra.iterrows():
        dsra_eop = row['DSRA EOP']
        if date(index.year, index.month, index.day) == date(2017, 1, 31):
            ''' first bop = first eop'''
            df_liquidity.set_value(index, 'DSRA BOP', dsra_eop)
        else:
            df_liquidity.set_value(index, 'DSRA BOP', dsra_bop)
        df_liquidity.set_value(index, 'DSRA EOP', dsra_eop)
        dsra_bop = dsra_eop
    return df_liquidity


def build_distributions(df_liquidity):
    df_dist = get_trial_balance('Distribution', 'total_debit', 'Distributions')
    for index, row in df_dist.iterrows():
        df_liquidity.set_value(index, 'Distributions', row['Distributions'])
    return df_liquidity


def build_other_cash_use(df_liquidity):
    ''' manual for now - need to identify in tb '''
    df_liquidity.set_value(date(2017,1,31), 'Other Cash Use', -4500000)
    df_liquidity.set_value(date(2017,3,31), 'Other Cash Use', -20790000)
    return df_liquidity


def build_change_work_cap(df_liquidity, forecast_start):
    for index, row in df_liquidity.iterrows():
        if date(index.year, index.month, index.day) < forecast_start:
            change_cash = row['Cash EOP'] - row['Cash BOP']
            activity_cash = row['EBITDA'] - row['Capex'] + row['Other Cash Use'] - row['TLB Int Exp'] - row['TLC Int Exp'] + row['TLB Change'] + row['DSRA Change'] - row['Distributions']
            change_work_cap = change_cash - activity_cash
        elif date(index.year, index.month, index.day) == utilities.calc_month_end(forecast_start, 'date'):
            change_work_cap = -df_liquidity.loc[:forecast_start,['Change Work Cap']].sum() - 20000000
        else:
            change_work_cap = 0
        df_liquidity.set_value(index, 'Change Work Cap', change_work_cap)
    return df_liquidity


def build_int_exp_actual(df_liquidity):
    ''' manaually input for now '''
    df_liquidity.set_value(date(2017,1,31), 'TLB Int Exp', 0)
    df_liquidity.set_value(date(2017,2,28), 'TLB Int Exp', 1015746.25)
    df_liquidity.set_value(date(2017,3,31), 'TLB Int Exp', 17067021.5)
    df_liquidity.set_value(date(2017,4,30), 'TLB Int Exp', 11012419.96)
    df_liquidity.set_value(date(2017,5,31), 'TLB Int Exp', 9763733.48)
    df_liquidity.set_value(date(2017,6,30), 'TLB Int Exp', 9248754.57)
    return df_liquidity


def build_tax_depreciation(forecast_end, scenario):
    entities = {'Gavin':[600000000, 0], 'Waterford':[550000000, 0], 'Lawrenceburg': [590000000, 0], 'Darby': [150000000, 0], 'HoldCo': [151622000, 0]}
    tax_dep = {2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0}

    current_year = 2017
    while current_year <= forecast_end.year:
        for entity in entities:
            old_tax_dep_rate = calc_tax_depreciation(entity, False)
            new_capex = get_new_capex(scenario, entity, current_year)
            entities[entity][0] += new_capex
            new_tax_dep_rate = calc_tax_depreciation(entity, True)
            tax_dep[current_year] += entities[entity][0] * new_tax_dep_rate + entities[entity][1] * old_tax_dep_rate
            entities[entity][1] += entities[entity][0]
            entities[entity][0] = 0
        current_year += 1
    return tax_dep


def build_est_tax_dist(df_liquidity, forecast_start, forecast_end, scenario):
    ''' manually assign oid and deferred finance charge (dfc) based on actuals '''
    oid = 410714.29 * 12
    dfc = 605031.96 * 12
    etr = 0.472684
    tax_split = {3:0.2623, 6:0.2459, 9:0.2459, 12:0.2459}  #use later
    tax_dep_all = build_tax_depreciation(forecast_end, scenario)
    current_year_begin = date(forecast_start.year, 1, 1)
    current_year_end = date(forecast_start.year, 12,31)
    while current_year_end <= forecast_end:
        ebitda = df_liquidity[current_year_begin:current_year_end]['EBITDA'].sum()
        int_exp_cash = df_liquidity[current_year_begin:current_year_end]['Int Exp'].sum()
        int_exp_tax = int_exp_cash + oid + dfc
        tax_dep = tax_dep_all[current_year_end.year]
        taxable_income = ebitda - int_exp_tax - tax_dep
        tax = max(taxable_income * etr, 0)
        if current_year_end == date(forecast_start.year, 12, 31):
            if forecast_start.month <= 3:
                df_liquidity.set_value(date(current_year_end.year, 3, 31),'Distributions', tax*0.2623)
                df_liquidity.set_value(date(current_year_end.year, 6, 30),'Distributions', tax*0.2459)
                df_liquidity.set_value(date(current_year_end.year, 9, 30),'Distributions', tax*0.2459)
                df_liquidity.set_value(date(current_year_end.year, 12, 31),'Distributions', tax*0.2459)
            if forecast_start.month <= 6:
                df_liquidity.set_value(date(current_year_end.year, 6, 30),'Distributions', tax*0.2459)
                df_liquidity.set_value(date(current_year_end.year, 9, 30),'Distributions', tax*0.2459)
                df_liquidity.set_value(date(current_year_end.year, 12, 31),'Distributions', tax*0.2459)
            elif forecast_start.month <= 9:
                df_liquidity.set_value(date(current_year_end.year, 9, 30),'Distributions', tax*0.2459)
                df_liquidity.set_value(date(current_year_end.year, 12, 31),'Distributions', tax*0.2459)
            else:
                df_liquidity.set_value(date(current_year_end.year, 12, 31),'Distributions', tax*0.2459)
        else:
            df_liquidity.set_value(date(current_year_end.year, 3, 31),'Distributions', tax*0.2623)
            df_liquidity.set_value(date(current_year_end.year, 6, 30),'Distributions', tax*0.2459)
            df_liquidity.set_value(date(current_year_end.year, 9, 30),'Distributions', tax*0.2459)
            df_liquidity.set_value(date(current_year_end.year, 12, 31),'Distributions', tax*0.2459)
        current_year_end = date(current_year_end.year+1, 12, 31)
        current_year_begin = date(current_year_begin.year+1, 1, 1)
    return df_liquidity


def build_tlb_bop_eop(df_liquidity):
    for index, row in df_liquidity.iterrows():
        if date(index.year, index.month, index.day) > date(2017, 1, 31):
            df_liquidity.set_value(index, 'Cash BOP', cash_bop)
            cash_eop = cash_bop + row['EBITDA'] - row['Capex'] + row['Other Cash Use'] - row['TLB Int Exp'] - row['TLC Int Exp'] \
                + row['TLB Change'] + row['DSRA Change'] - row['Distributions'] + row['Change Work Cap']
            df_liquidity.set_value(index, 'Cash EOP', cash_eop)
        cash_bop = row['Cash EOP']
    return df_liquidity


def search_for_prepay(df_liquidity, forecast_start, min_cash):
    for index, row in df_liquidity.iterrows():
        if date(index.year, index.month, index.day) >= forecast_start:
            if index.month % 3 == 0:
                if row['Cash EOP'] > min_cash:
                    prepay = math.ceil(row['Cash EOP'] - min_cash)
                    df_liquidity.set_value(index, 'TLB Prepay', prepay)
                    break
    return df_liquidity


def recalc_tlb_amort(df, forecast_start):
    for index, row in df.iterrows():
        if date(index.year, index.month, index.day) >= forecast_start:
            if index.month % 3 == 0:
                required_amort = calc_required_amort(df, 'Lightstone TLB', index)
                cum_prepay = df.loc[:index]['TLB Prepay'].sum()
                cum_amort = df.loc[:index]['TLB Amort'].sum()
                tlb_amort = max(min(1625000000*0.0025,required_amort-cum_prepay-cum_amort),0)
                df.set_value(index, ['TLB Amort'], tlb_amort)
    return df


def recalc_tlb_int_exp(df_liquidity, forecast_start):
    for index, row in df_liquidity.iterrows():
        if date(index.year, index.month, index.day) >= forecast_start:
            swap_int_exp = row['Swap Avg Daily Balance'] * row['Swap Fix Rate']
            float_int_exp = (row['TLB BOP'] - row['Swap Avg Daily Balance']) * row['TLB Float Rate']
            df_liquidity.set_value(index, 'TLB Int Exp', (swap_int_exp + float_int_exp) * index.day/360)
    return df_liquidity


def recalc_tlb(df_liquidity, forecast_start):
    for index, row in df_liquidity.iterrows():
        if date(index.year, index.month, index.day) >= forecast_start:
            tlb_eop = tlb_bop + row['TLB Addl Borrow'] - row['TLB Amort'] - row['TLB Prepay']
            df_liquidity.set_value(index, 'TLB EOP', tlb_eop)
        tlb_bop = row['TLB EOP']
    df_liquidity['TLB Change'] = df_liquidity['TLB EOP'].diff()
    df_liquidity.set_value(date(2017, 1,31), 'TLB Change', 0)   #no diff for first entry
    return df_liquidity


def write_report(df_liquidity):
    dft = df_liquidity[:date(2017, 12,31)].transpose()

    column_labels = df_liquidity[:date(2017,12,31)].index
    row_labels = {'Cash BOP':['Beginning cash', 5, 1],
                'EBITDA':['EBITDA', 6, 1],
                'Capex':['Capex', 7, -1],
                'Change Work Cap':['Change in working capital', 8, 1],
                'Other Cash Use':['Other cash use', 9, 1],
                'Int Exp': ['Cash interest expense', 10, -1],
                'Revolver Change': ['Change in Revolver', 11, 1],
                'TLB Change': ['Change in TLB', 12, 1],
                'DSRA Change': ['Change in DSRA', 13, 1],
                'Distributions': ['Distributions', 14, -1],
                'Cash EOP': ['Ending cash', 15, 1]}

    wb = opx.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width=22.0
    column_header_format = NamedStyle(name='datetime', number_format='mmm-yy')
    data_format = NamedStyle(name='data', number_format='#,##0.0_);(#,##0.0)')
    for row in dataframe_to_rows(dft, index=True, header=True):
        for index, col in  enumerate(column_labels):
            ws.cell(row=4, column=index+2, value=col)
            ws.cell(row=4, column=index+2).style = column_header_format
        if row[0] in row_labels:
            ws.cell(row=row_labels[row[0]][1], column=1, value=row_labels[row[0]][0])
            for i in range(12):
                ws.cell(row=row_labels[row[0]][1], column=i+2, value=row[i+1]/1000000*row_labels[row[0]][2])
                ws.cell(row=row_labels[row[0]][1], column=i+2).style = data_format

    wb.save('pandas_xl.xlsx')


''' ===========================  MAIN =========================== '''
valuation_date = date(2017,6,28)
forecast_start = date(2017,7,1)
forecast_end = date(2021,12,31)
scenario_start = date(2017,1,30)
scenario_debt_inception = 'Actuals'
scenario = '2017 June AMR'
company = 'Lightstone'
entity = 'Holdco'
instrument_id = 'Lightstone TLB'
min_cash = 50000000

periods = pd.date_range(scenario_start, forecast_end, freq='M')
df_liquidity = pd.DataFrame(periods, columns = ['period'])
df_liquidity.set_index(['period'], inplace=True)
df_liquidity['TLB BOP'] = 0
df_liquidity['TLB Addl Borrow'] = 0
df_liquidity['TLB Amort'] = 0
df_liquidity['TLB Prepay'] = 0
df_liquidity['TLB EOP'] = 0
df_liquidity['TLB Change'] = 0
df_liquidity['TLC BOP'] = 0
df_liquidity['TLC Addl Borrow'] = 0
df_liquidity['TLC Amort'] = 0
df_liquidity['TLC Prepay'] = 0
df_liquidity['TLC EOP'] = 0
df_liquidity['Revolver Change'] = 0
df_liquidity['LIBOR'] = 0
df_liquidity['TLB Float Rate'] = 0
df_liquidity['Swap Fix Rate'] = 0
df_liquidity['Swap Avg Daily Balance'] = 0
df_liquidity['TLB Int Exp'] = 0
df_liquidity['TLC Float Rate'] = 0
df_liquidity['TLC Int Exp'] = 0
df_liquidity['Int Exp'] = 0
df_liquidity['EBITDA'] = 0
df_liquidity['Tax Depreciation'] = 0
df_liquidity['Taxable Interest'] = 0
df_liquidity['Eff Tax Rate'] = 0
df_liquidity['Est Tax Dist'] = 0
df_liquidity['Cash BOP'] = 0
df_liquidity['Capex'] = 0
df_liquidity['Change Work Cap'] = 0
df_liquidity['Other Cash Use'] = 0
df_liquidity['Distributions'] = 0
df_liquidity['Cash EOP'] = 0
df_liquidity['DSRA BOP'] = 0
df_liquidity['DSRA EOP'] = 0
df_liquidity['DSRA Change'] = 0
df_liquidity['Excess Cash'] = 0

df_liquidity = build_debt(df_liquidity, forecast_start, forecast_end)
df_liquidity['TLB Change'] = df_liquidity['TLB EOP'].diff()
df_liquidity.set_value(date(2017, 1,31), 'TLB Change', 0)   #no diff for first entry
df_libor_all = get_all_libor(valuation_date, 'LIBOR-1MO')
df_liquidity = build_libor(df_liquidity, df_libor_all, forecast_start, valuation_date)
df_liquidity['TLB Float Rate'] = df_liquidity['LIBOR'] + 0.045
df_swap = build_daily_swap(valuation_date, forecast_start, forecast_end)
df_liquidity = build_swap_rate(df_liquidity, df_swap, forecast_start)
df_liquidity = build_tlb_int_exp(df_liquidity, forecast_start)
df_liquidity = build_int_exp_actual(df_liquidity)
df_liquidity.loc[forecast_start:,['TLC Float Rate']] = df_liquidity['LIBOR'] + 0.045
df_liquidity.loc[forecast_start:,['TLC Int Exp']] = df_liquidity['TLC Float Rate'] * df_liquidity['TLC BOP'] * df_liquidity.index.day / 360
df_liquidity['Int Exp'] = df_liquidity['TLB Int Exp'] + df_liquidity['TLC Int Exp']
df_liquidity = build_ebitda(df_liquidity, scenario, company)
df_liquidity = build_capex(df_liquidity, scenario, company)

df_cash = get_cash()
df_liquidity = build_cash_actuals(df_liquidity, df_cash)
df_liquidity = build_dsra(df_liquidity)
df_liquidity['DSRA Change'] = -df_liquidity['DSRA EOP'].diff()
df_liquidity.set_value(date(2017, 1,31), 'DSRA Change', 0)   #no diff for first entry
df_liquidity = build_distributions(df_liquidity)
df_liquidity = build_other_cash_use(df_liquidity)

df_liquidity = build_change_work_cap(df_liquidity, forecast_start)

df_liquidity = build_est_tax_dist(df_liquidity, forecast_start, forecast_end, scenario)

'''iterative search '''
for i in range(20):
    df_liquidity = build_tlb_bop_eop(df_liquidity)
    df_liquidity = search_for_prepay(df_liquidity, forecast_start, min_cash)
    df_liquidity = recalc_tlb_amort(df_liquidity, forecast_start)
    df_liquidity = recalc_tlb(df_liquidity, forecast_start)
    df_liquidity = recalc_tlb_int_exp(df_liquidity, forecast_start)
    df_liquidity = build_est_tax_dist(df_liquidity, forecast_start, forecast_end, scenario)


#df_cash.to_csv('df_cash.csv')
#df_swap.to_csv('df_swap.csv')
df_liquidity.to_csv('df_liquidity.csv')

write_report(df_liquidity)
