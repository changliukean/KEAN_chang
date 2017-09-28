''' TODO: Add test to see if spreadsheet currently open '''


import mysql.connector
from mysql.connector import errorcode

import pandas as pd

import openpyxl as opx
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle

from calendar import monthrange
from datetime import date, datetime
import time

import os
import sys

path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
path_utlities = path + '/utilities/'
sys.path.insert(0, path_utlities)
import utilities

def get_hedges(company, entity, date_start, date_end):

    date_start_str = datetime.strftime(date_start, '%Y-%m-%d')
    date_end_str = datetime.strftime(date_end, '%Y-%m-%d')

    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')

    query = "SELECT * FROM hedges WHERE company = \'" + company + \
            "\' AND entity = \'" + entity + "\' AND date_start <= \'" \
            + date_end_str + "\' AND date_end >= \'" + date_start_str + "\' " \
            + "ORDER BY date_start, tid_kindle"
    df_hedges = pd.read_sql(query, cnx)

    cnx.close

    return df_hedges


def get_all_commodity_prices(valuation_date):

    val_date_str = datetime.strftime(valuation_date, '%Y-%m-%d')
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')

    query = "SELECT instrument_id, period, price FROM prices WHERE valuation_date = \'" + val_date_str + "\'"
    df_prices = pd.read_sql(query, cnx)

    cnx.close

    return df_prices


def get_specific_price(df, instrument_id, period):
    ''' pass the df from get_all_commodity_prices - order is instrument_id, period, price '''
    ''' cheat and allow broad range of instrument_ids to be passed then cleaned up '''
    instrument_id = 'AD Hub ATC - Monthly Strip' if instrument_id == 'AD Hub DA ATC' else instrument_id
    instrument_id = 'AD Hub On Peak - Monthly Strip' if instrument_id == 'AD Hub DA Peak' else instrument_id
    instrument_id = 'AD Hub On Peak - Monthly Strip' if instrument_id == 'AD Hub RT Peak' else instrument_id
    instrument_id = 'Tetco ELA - Monthly Strip' if instrument_id == 'Tetco ELA' else instrument_id
    #instrument_id = 'TGT Zone 1 - Monthly Strip' if instrument_id == '' else instrument_id

    price = df.loc[(df['instrument_id']==instrument_id) & (df['period']==period)]

    price = 0 if price.empty else price.iloc[0]['price']

    return price


def calc_monthly_cashflow(price_receive, price_pay, notional, frequency, period):
    days = monthrange(period.year, period.month)[1]
    periods = days if frequency == 'Daily' else days * 24
    cashflow = (price_receive - price_pay) * notional * periods

    return cashflow


def load_hedges_financials(scenario, company, entity, account, period, value):
    cnx = mysql.connector.connect(host = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com',
                                user = 'Andrew', password = 'Kindle01', database = 'poc_test_cl')
    cursor = cnx.cursor()
    query = ("INSERT INTO financials(scenario, company, entity, account, period, value) "
            "VALUES(%s, %s, %s, %s, %s, %s)")
    try:
        cursor.execute(query, [scenario, company, entity, account, period, value])
    except mysql.connector.Error as err:
        print(err.msg)
    cnx.commit()
    cursor.close()
    cnx.close()


''' Need to develop flexible means of accomodating different instrument_id '''
'''    e.g. AD Hub DA ATC vs AD Hub RT ATC vs AD Hub ATC - Monthly Strip '''

valuation_date = date(2017, 6, 28)
forecast_start = date(2017, 7,1)
forecast_end = date(2021, 12, 31)

df_hedges = get_hedges('Lightstone', 'HoldCo', forecast_start, forecast_end)
df_all_prices = get_all_commodity_prices(valuation_date)

excel_template = 'c:/users/agood/projects/kean_proj/templates/hedge detail.xlsx'
wb = opx.load_workbook(filename = excel_template)
ws = wb.active

image =  Image('c:/users/agood/projects/kean_proj/graphics/Lightstone_Logo.jpg')
ws.add_image(image, 'A1')
ws.row_dimensions[1].height = 54

''' clear main body of worksheet '''
for row in ws.iter_rows(row_offset=4):
    for cell in row:
        cell.value = None

loop_rows = 1
remaining_column = 8 + (13 - forecast_start.month) + 1
data_format = NamedStyle(name='data', number_format='#,##0_);(#,##0)')
df_hedges_relevant = df_hedges[df_hedges['trade_date']<forecast_start]
df_hedges_relevant = df_hedges_relevant.reset_index(drop=True)
for i, hedge in df_hedges_relevant.iterrows():
    ws.cell(row=i+5, column=1, value=hedge.loc['tid_kindle'])
    ws.cell(row=i+5, column=2, value=hedge.loc['counterparty'])
    instrument = hedge.loc['receive_index'] + "/" + hedge.loc['pay_index']
    ws.cell(row=i+5, column=3, value=instrument)
    ws.cell(row=i+5, column=4, value=hedge.loc['notional'])
    fixed_price = hedge.loc['receive_value'] if hedge.loc['receive_index'] == 'fixed' else hedge.loc['pay_value']
    ws.cell(row=i+5, column=5, value=fixed_price)
    ws.cell(row=i+5, column=6, value=hedge.loc['frequency'])
    ws.cell(row=i+5, column=7, value=hedge.loc['date_start'])
    ws.cell(row=i+5, column=8, value=hedge.loc['date_end'])


    """show monthly results for remainder of forecast year, then sum remaining into one column"""
    loop_date = utilities.calc_month_end(forecast_start, 'date')
    month_offset = 0
    remaining_cashflow = 0

    while loop_date <= forecast_end:
        if hedge.loc['date_start'] <= loop_date and hedge.loc['date_end'] >= loop_date:
            price_receive = hedge.loc['receive_value'] if hedge.loc['receive_index'] == 'fixed' else get_specific_price(df_all_prices, hedge.loc['receive_index'], loop_date)
            price_pay = hedge.loc['pay_value'] if hedge.loc['pay_index'] == 'fixed' else get_specific_price(df_all_prices, hedge.loc['pay_index'], loop_date)
            monthly_cashflow = calc_monthly_cashflow(price_receive, price_pay, hedge.loc['notional'], hedge.loc['frequency'], loop_date)
            if loop_date.year == forecast_start.year:
                ws.cell(row=i+5, column=9+month_offset, value=monthly_cashflow)
                ws.cell(row=i+5, column=9+month_offset).style = data_format
            else:
                remaining_cashflow += monthly_cashflow
        if loop_date > hedge.loc['date_end']:
            ws.cell(row=i+5, column=remaining_column, value=remaining_cashflow)
            ws.cell(row=i+5, column=remaining_column).style = data_format
            break

        month_offset += 1
        loop_date = utilities.calc_next_month_end(loop_date, 'date')

    loop_rows += 1


''' add totals row '''
ws.cell(row=loop_rows+4, column=1, value='Total')
for i in range(9, remaining_column+1):
    formula = '=SUM(' + chr(96+i) + str(1+4) + ':' + chr(96+i) + str(loop_rows+4-1) + ')'
    ws.cell(row=loop_rows+4, column=i, value=formula)
    ws.cell(row=loop_rows+4, column=i).style = data_format

''' Add bold font and grey fill to total row '''
for i in range(1, remaining_column+1):
    current_cell = chr(96+i) + str(loop_rows+4)
    ws[current_cell].fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    ws[current_cell].font = Font(bold=True)

''' Add underline to report title '''
for i in range(1, remaining_column+1):
    current_cell = chr(96+i) + str(1)
    ws[current_cell].border = Border(bottom=Side(style='thick'))

label = 'Pricing Date: ' +  valuation_date.strftime('%m/%d/%Y')
ws.cell(row=1, column=remaining_column, value=label)

excel_report = 'c:/users/agood/projects/kean_proj/reports/hedge detail.xlsx'
wb.save(filename=excel_report)


''' write results to financials '''

loop_date = utilities.calc_month_end(forecast_start, 'date')
while loop_date <= forecast_end:
    sum_cashflow = 0.0
    for i, hedge in df_hedges.iterrows():
        if hedge.loc['date_start'] <= loop_date and hedge.loc['date_end'] >= loop_date:
            price_receive = hedge.loc['receive_value'] if hedge.loc['receive_index'] == 'fixed' else get_specific_price(df_all_prices, hedge.loc['receive_index'], loop_date)
            price_pay = hedge.loc['pay_value'] if hedge.loc['pay_index'] == 'fixed' else get_specific_price(df_all_prices, hedge.loc['pay_index'], loop_date)
            monthly_cashflow = calc_monthly_cashflow(price_receive, price_pay, hedge.loc['notional'], hedge.loc['frequency'], loop_date)
            sum_cashflow += monthly_cashflow

    #load_hedges_financials('2017 June AMR', 'Lightstone', 'HoldCo', 'Hedge P&L', loop_date, '%.2f' % sum_cashflow)
    loop_date = utilities.calc_next_month_end(loop_date, 'date')
